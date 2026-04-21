"""Tests for seller exclusion patterns in report generation."""
import io

import pytest
from decimal import Decimal
from openpyxl import load_workbook

from app.services.csv_generator import CSVGenerator


@pytest.mark.unit
def test_seller_display_excluded_metroshoe_variants():
    patterns = ["metroshoe"]
    assert CSVGenerator.seller_display_excluded("MetroShoe", patterns)
    assert CSVGenerator.seller_display_excluded("metroshoe", patterns)
    assert CSVGenerator.seller_display_excluded("Metroshoe Warehouse", patterns)
    assert CSVGenerator.seller_display_excluded("metroshoe warehouse", patterns)
    assert CSVGenerator.seller_display_excluded("Metro Shoe Warehouse", patterns)
    assert not CSVGenerator.seller_display_excluded("Other Vendor Inc", patterns)
    assert not CSVGenerator.seller_display_excluded("N/A", patterns)


@pytest.mark.unit
def test_seller_display_excluded_empty_patterns():
    assert not CSVGenerator.seller_display_excluded("MetroShoe", None)
    assert not CSVGenerator.seller_display_excluded("MetroShoe", [])


@pytest.mark.unit
def test_comprehensive_report_skips_excluded_seller_row():
    # Minimal keepa: one offer below MAP from "MetroShoe"
    keepa = {
        "products": [
            {
                "asin": "B00TEST",
                "title": "T",
                "brand": "B",
                "current_sellers": [
                    {
                        "sellerId": "M1",
                        "sellerName": "MetroShoe Warehouse",
                        "price": 1000,
                        "isFBA": False,
                        "condition": "New",
                    }
                ],
                "offers": [],
            }
        ]
    }
    items = [{"upc": "123", "keepa_data": keepa, "status": "completed"}]
    excel_bytes, count = CSVGenerator.generate_comprehensive_report_csv(
        items,
        {"123": Decimal("50.00")},
        seller_name_map={},
        excluded_seller_substrings=["metroshoe"],
    )
    assert count == 0
    assert len(excel_bytes) > 0


@pytest.mark.unit
def test_comprehensive_report_keeps_non_excluded_seller():
    keepa = {
        "products": [
            {
                "asin": "B00TEST",
                "title": "T",
                "brand": "B",
                "stats": {
                    "buyBoxSellerId": "X1",
                    "buyBoxPrice": 1000,
                },
                "current_sellers": [
                    {
                        "sellerId": "X1",
                        "sellerName": "Acme Retail",
                        "price": 1000,
                        "isFBA": False,
                        "condition": "New",
                    }
                ],
                "offers": [],
            }
        ]
    }
    items = [{"upc": "123", "keepa_data": keepa, "status": "completed"}]
    _, count = CSVGenerator.generate_comprehensive_report_csv(
        items,
        {"123": Decimal("50.00")},
        seller_name_map={},
        excluded_seller_substrings=["metroshoe"],
    )
    assert count == 1


@pytest.mark.unit
def test_comprehensive_report_one_row_per_upc_prefers_buy_box():
    """Multiple sellers below MAP: single row; buy-box seller wins over cheaper offers."""
    keepa = {
        "products": [
            {
                "asin": "B00TEST",
                "title": "T",
                "brand": "B",
                "stats": {
                    "buyBoxSellerId": "WINNER",
                    "buyBoxPrice": 2500,
                },
                "current_sellers": [
                    {
                        "sellerId": "WINNER",
                        "sellerName": "Winner Co",
                        "price": 2500,
                        "isFBA": False,
                        "condition": "New",
                    },
                    {
                        "sellerId": "CHEAP",
                        "sellerName": "Cheapo",
                        "price": 2000,
                        "isFBA": False,
                        "condition": "New",
                    },
                ],
                "offers": [],
            }
        ]
    }
    items = [{"upc": "123", "keepa_data": keepa, "status": "completed"}]
    excel_bytes, count = CSVGenerator.generate_comprehensive_report_csv(
        items,
        {"123": Decimal("50.00")},
        seller_name_map={},
        excluded_seller_substrings=[],
    )
    assert count == 1
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    seller_col = headers.index("Seller") + 1
    price_col = headers.index("Seller Offer Price") + 1
    assert ws.cell(row=2, column=seller_col).value == "Winner Co"
    assert ws.cell(row=2, column=price_col).value == "$25.00"


@pytest.mark.unit
def test_comprehensive_report_does_not_flag_non_buy_box_seller():
    """Even if another seller is below MAP, no row is added unless buy-box winner is below MAP."""
    keepa = {
        "products": [
            {
                "asin": "B00TEST",
                "title": "T",
                "brand": "B",
                "stats": {
                    "buyBoxSellerId": "WINNER",
                    "buyBoxPrice": 5500,
                },
                "current_sellers": [
                    {
                        "sellerId": "WINNER",
                        "sellerName": "Winner Co",
                        "price": 5500,
                        "isFBA": False,
                        "condition": "New",
                    },
                    {
                        "sellerId": "CHEAP",
                        "sellerName": "Cheapo",
                        "price": 2000,
                        "isFBA": False,
                        "condition": "New",
                    },
                ],
                "offers": [],
            }
        ]
    }
    items = [{"upc": "123", "keepa_data": keepa, "status": "completed"}]
    _, count = CSVGenerator.generate_comprehensive_report_csv(
        items,
        {"123": Decimal("50.00")},
        seller_name_map={},
        excluded_seller_substrings=[],
    )
    assert count == 0


@pytest.mark.unit
def test_comprehensive_report_dedupes_duplicate_upc_last_wins():
    keepa_first = {
        "products": [
            {
                "asin": "BOLD",
                "title": "Old",
                "brand": "B",
                "stats": {
                    "buyBoxSellerId": "X1",
                    "buyBoxPrice": 1000,
                },
                "current_sellers": [
                    {
                        "sellerId": "X1",
                        "sellerName": "First Seller",
                        "price": 1000,
                        "isFBA": False,
                        "condition": "New",
                    }
                ],
                "offers": [],
            }
        ]
    }
    keepa_second = {
        "products": [
            {
                "asin": "BNEW",
                "title": "New Title",
                "brand": "B",
                "stats": {
                    "buyBoxSellerId": "Y1",
                    "buyBoxPrice": 1000,
                },
                "current_sellers": [
                    {
                        "sellerId": "Y1",
                        "sellerName": "Second Seller",
                        "price": 1000,
                        "isFBA": False,
                        "condition": "New",
                    }
                ],
                "offers": [],
            }
        ]
    }
    items = [
        {"upc": "123", "keepa_data": keepa_first, "status": "completed"},
        {"upc": "123", "keepa_data": keepa_second, "status": "completed"},
    ]
    excel_bytes, count = CSVGenerator.generate_comprehensive_report_csv(
        items,
        {"123": Decimal("50.00")},
        seller_name_map={},
        excluded_seller_substrings=[],
    )
    assert count == 1
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    title_col = headers.index("Product Title") + 1
    assert ws.cell(row=2, column=title_col).value == "New Title"


@pytest.mark.unit
def test_extract_keepa_product_data_matches_buybox_seller_id_across_types():
    """Seller id matching should work even when buyBoxSellerId type differs from sellerId type."""
    keepa = {
        "products": [
            {
                "asin": "B00TEST",
                "title": "T",
                "brand": "B",
                "stats": {
                    "buyBoxSellerId": 12345,
                    "buyBoxPrice": None,
                },
                "current_sellers": [
                    {
                        "sellerId": "12345",
                        "sellerName": "Correct Seller",
                        "price": 2500,
                        "isFBA": False,
                        "condition": "New",
                    },
                    {
                        "sellerId": "99999",
                        "sellerName": "Wrong Seller",
                        "price": 2400,
                        "isFBA": True,
                        "condition": "New",
                    },
                ],
                "offers": [],
            }
        ]
    }
    product = CSVGenerator.extract_keepa_product_data(keepa)
    assert product["buy_box_seller_id"] == "12345"
    assert product["buy_box_seller_name"] == "Correct Seller"
    assert product["buy_box_price"] == 25.0
