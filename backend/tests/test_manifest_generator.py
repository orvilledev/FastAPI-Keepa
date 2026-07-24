"""Tests for Manifest Generator packing-sheet → STA zip conversion."""
from __future__ import annotations

import zipfile
from datetime import date
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

from app.services.manifest_generator import (
    TEMPLATE_SHEET_NAME,
    build_manifest_zip,
    format_manifest_date,
    normalize_vendor_label,
)

FIXTURES = Path(__file__).resolve().parent / "fixtures" / "manifest_generator"
SAMPLE_INPUT = FIXTURES / "input.xlsx"
SAMPLE_ZIP = FIXTURES / "expected.zip"


pytestmark = pytest.mark.skipif(
    not SAMPLE_INPUT.exists() or not SAMPLE_ZIP.exists(),
    reason="Manifest Generator fixtures not present",
)


def _template_items(xlsx_bytes: bytes) -> list[tuple[str, int]]:
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    try:
        assert TEMPLATE_SHEET_NAME in wb.sheetnames
        assert wb.sheetnames == [
            "Instructions",
            "Data definitions",
            TEMPLATE_SHEET_NAME,
            "Create workflow \u2013 example",
        ]
        ws = wb[TEMPLATE_SHEET_NAME]
        items: list[tuple[str, int]] = []
        for row in range(9, (ws.max_row or 8) + 1):
            sku = ws.cell(row, 1).value
            qty = ws.cell(row, 2).value
            if sku is None:
                continue
            items.append((str(sku), int(qty)))
        return items
    finally:
        wb.close()


def test_format_manifest_date_no_padding():
    assert format_manifest_date(date(2026, 7, 24)) == "7.24.26"
    assert format_manifest_date(date(2026, 12, 5)) == "12.5.26"


def test_normalize_vendor_place_suffix():
    assert normalize_vendor_label("FTWSWL PLACE") == "FTWSWL"
    assert normalize_vendor_label("FTWSWL") == "FTWSWL"
    assert normalize_vendor_label("ZSE") == "ZSE"


def test_sample_input_matches_sample_zip():
    result = build_manifest_zip(SAMPLE_INPUT.read_bytes(), when=date(2026, 7, 24))
    assert result.zip_filename == "FTWSWL FBA Manifests 7.24.26.zip"
    assert result.file_count == 4
    assert result.primary_vendor == "FTWSWL"

    with zipfile.ZipFile(BytesIO(result.zip_bytes)) as got_zip, zipfile.ZipFile(SAMPLE_ZIP) as exp_zip:
        assert sorted(got_zip.namelist()) == sorted(exp_zip.namelist())
        for name in sorted(exp_zip.namelist()):
            got_items = _template_items(got_zip.read(name))
            exp_items = _template_items(exp_zip.read(name))
            assert got_items == exp_items, f"Mismatch in {name}"
