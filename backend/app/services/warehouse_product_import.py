"""Parse PRODUCTS sheet / CSV for warehouse product catalog import."""
from __future__ import annotations

import csv
import io
import logging
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from app.repositories.warehouse_product_repository import normalize_upc_key

logger = logging.getLogger(__name__)

_HEADER_ALIASES = {
    "upc": "upc",
    "fnsku": "fnsku",
    "style name": "style_name",
    "style_name": "style_name",
    "condition": "condition",
}


def _normalize_header(cell: Any) -> Optional[str]:
    if cell is None:
        return None
    key = str(cell).strip().lower()
    return _HEADER_ALIASES.get(key)


def _parse_row(mapping: Dict[str, int], values: Tuple[Any, ...]) -> Optional[Dict[str, str]]:
    def cell(name: str) -> str:
        idx = mapping.get(name)
        if idx is None or idx >= len(values):
            return ""
        raw = values[idx]
        if raw is None:
            return ""
        return str(raw).strip()

    upc = normalize_upc_key(cell("upc"))
    fnsku = cell("fnsku")
    if not upc or not fnsku:
        return None
    return {
        "upc": upc,
        "fnsku": fnsku,
        "style_name": cell("style_name"),
        "condition": cell("condition") or "New",
    }


def _header_mapping_from_row(headers: Tuple[Any, ...]) -> Optional[Dict[str, int]]:
    mapping: Dict[str, int] = {}
    for idx, header in enumerate(headers):
        normalized = _normalize_header(header)
        if normalized:
            mapping[normalized] = idx
    if "upc" not in mapping or "fnsku" not in mapping:
        return None
    return mapping


def parse_products_csv(content: bytes) -> Tuple[List[Dict[str, str]], int]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], 0
    mapping = _header_mapping_from_row(tuple(rows[0]))
    if not mapping:
        raise ValueError(
            'CSV must include headers "UPC" and "fnsku" (and optionally "STYLE NAME", "Condition").'
        )
    valid: List[Dict[str, str]] = []
    invalid = 0
    for row in rows[1:]:
        if not any(str(c).strip() for c in row):
            continue
        parsed = _parse_row(mapping, tuple(row))
        if parsed:
            valid.append(parsed)
        else:
            invalid += 1
    return valid, invalid


def _sheet_rows_from_workbook(filename: str, content: bytes) -> List[Tuple[Any, ...]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = None
        for name in wb.sheetnames:
            if name.strip().upper() == "PRODUCTS":
                sheet = wb[name]
                break
        if sheet is None:
            sheet = wb[wb.sheetnames[0]]
        return [tuple(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        wb.close()


def parse_products_spreadsheet(filename: str, content: bytes) -> Tuple[List[Dict[str, str]], int]:
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        return parse_products_csv(content)

    if not lower.endswith((".xlsx", ".xlsm", ".xls")):
        raise ValueError("Upload a .csv, .xlsx, or .xlsm file (PRODUCTS sheet).")

    rows = _sheet_rows_from_workbook(filename, content)
    if not rows:
        return [], 0

    header_idx = 0
    mapping: Optional[Dict[str, int]] = None
    for idx, row in enumerate(rows[:20]):
        candidate = _header_mapping_from_row(row)
        if candidate:
            mapping = candidate
            header_idx = idx
            break
    if not mapping:
        raise ValueError(
            'Spreadsheet must include columns "UPC" and "fnsku" (PRODUCTS sheet).'
        )

    valid: List[Dict[str, str]] = []
    invalid = 0
    for row in rows[header_idx + 1 :]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        parsed = _parse_row(mapping, row)
        if parsed:
            valid.append(parsed)
        else:
            invalid += 1
    return valid, invalid


def dedupe_by_upc(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """Last row wins for duplicate UPCs within one import file."""
    seen: Dict[str, Dict[str, str]] = {}
    for row in rows:
        seen[row["upc"]] = row
    return list(seen.values())
