"""DNK AllInventory — split PMSH01 available-inventory export into per-PO sheets."""
from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

MAIN_HEADERS: Tuple[str, ...] = (
    "Sell to Customer No ",
    "Description",
    "Dansko Order No",
    "PONumber",
    "Cross Reference No ",
    "Remaining Order Qty",
    "Available Qty",
    "Req Shipment Date",
    "Item No",
    "Variant Code",
    "Length",
    "Width",
    "Height",
    "Weight",
)

PO_HEADERS: Tuple[str, ...] = (
    "PONumber",
    "Cross Reference No ",
    "Remaining Order Qty",
    "Available Qty",
)

# Canonical header key -> accepted aliases (normalized)
_REQUIRED = ("ponumber", "cross reference no", "remaining order qty", "available qty")

_MAIN_COL_WIDTHS = {
    "A": 18.0,
    "B": 30.0,
    "C": 15.86,
    "D": 16.43,
    "E": 17.57,
    "F": 18.29,
    "G": 13.71,
    "H": 16.71,
    "I": 13.71,
    "J": 12.0,
    "K": 10.0,
    "L": 10.0,
    "M": 10.0,
    "N": 10.0,
}

_PO_COL_WIDTHS = {
    "A": 16.43,
    "B": 17.57,
    "C": 18.29,
    "D": 13.71,
}

_HEADER_FONT_SEGOE = Font(name="Segoe UI", size=8, bold=True)
_HEADER_FONT_ARIAL = Font(name="Arial", size=8, bold=True)
_DATA_FONT = Font(name="Calibri", size=11)

# Column index (0-based) -> header font style on main sheet (matches source export)
_MAIN_HEADER_FONTS = {
    0: _HEADER_FONT_SEGOE,
    1: _HEADER_FONT_SEGOE,
    2: _HEADER_FONT_ARIAL,
    3: _HEADER_FONT_ARIAL,
    4: _HEADER_FONT_SEGOE,
    5: _HEADER_FONT_ARIAL,
    6: _HEADER_FONT_ARIAL,
    7: _HEADER_FONT_ARIAL,
    8: _HEADER_FONT_ARIAL,
    9: _HEADER_FONT_ARIAL,
    10: _HEADER_FONT_ARIAL,
    11: _HEADER_FONT_ARIAL,
    12: _HEADER_FONT_ARIAL,
    13: _HEADER_FONT_ARIAL,
}

_CENTER_COLS = {5, 6, 10, 11, 12, 13}  # Remaining, Available, L/W/H/Weight


class DnkAllInventoryError(ValueError):
    """Raised when the uploaded inventory workbook cannot be processed."""


@dataclass
class DnkAllInventoryResult:
    file_bytes: bytes
    filename: str
    main_sheet_name: str
    row_count: int
    po_count: int
    available_gt_zero: int
    date_stamp: str
    po_names: List[str] = field(default_factory=list)


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip().lower()


def _sheet_date_stamp(now: Optional[datetime] = None) -> str:
    """Match existing product naming: M.D.YY (no zero-padding)."""
    dt = now or datetime.now(ZoneInfo("America/New_York"))
    return f"{dt.month}.{dt.day}.{str(dt.year)[-2:]}"


def _safe_sheet_title(name: str) -> str:
    cleaned = re.sub(r'[\\/*?:\[\]]', "_", name).strip() or "Sheet"
    return cleaned[:31]


def _as_str(value: Any, *, strip: bool = False) -> str:
    """Convert cell values to text without forcing numeric scientific notation.

    By default trailing/leading spaces are preserved (matches source export descriptions).
    Pass strip=True for identifier fields such as PONumber.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value:  # NaN
            return ""
        if value == int(value) and abs(value) < 1e15:
            return str(int(value))
        text = str(value)
        if "." in text:
            text = text.rstrip("0").rstrip(".")
        return text
    text = str(value)
    if strip:
        text = text.strip()
    if re.fullmatch(r"-?\d+\.0+", text.strip()):
        return text.strip().split(".", 1)[0]
    return text


def _as_int(value: Any, *, field_name: str, default: int = 0) -> int:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        raise DnkAllInventoryError(f"{field_name} must be a number, got boolean.")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value != value:
            return default
        if abs(value - round(value)) > 1e-9:
            raise DnkAllInventoryError(f"{field_name} must be a whole number, got {value!r}.")
        return int(round(value))
    text = str(value).strip().replace(",", "")
    if not text:
        return default
    try:
        if "." in text:
            num = float(text)
            if abs(num - round(num)) > 1e-9:
                raise DnkAllInventoryError(f"{field_name} must be a whole number, got {value!r}.")
            return int(round(num))
        return int(text)
    except (TypeError, ValueError) as exc:
        raise DnkAllInventoryError(f"{field_name} must be a number, got {value!r}.") from exc


def _as_number(value: Any, *, field_name: str) -> Any:
    """Return int for whole numbers, float otherwise; blank -> 0."""
    if value is None or value == "":
        return 0
    if isinstance(value, bool):
        raise DnkAllInventoryError(f"{field_name} must be a number, got boolean.")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value != value:
            return 0
        if value == int(value):
            return int(value)
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0
    try:
        num = float(text)
    except ValueError as exc:
        raise DnkAllInventoryError(f"{field_name} must be a number, got {value!r}.") from exc
    if num == int(num):
        return int(num)
    return float(num)


def _as_upc_int(value: Any, *, field_name: str) -> int:
    """PO sheets store Cross Reference No as a numeric (int) cell."""
    text = _as_str(value, strip=True)
    if not text:
        raise DnkAllInventoryError(f"{field_name} is required.")
    if not re.fullmatch(r"\d+", text):
        try:
            return _as_int(value, field_name=field_name)
        except DnkAllInventoryError as exc:
            raise DnkAllInventoryError(f"{field_name} must be numeric, got {value!r}.") from exc
    return int(text)


def _find_inventory_sheet(wb) -> Worksheet:
    """Prefer a sheet whose headers include the required inventory columns."""
    candidates: List[Worksheet] = []
    for ws in wb.worksheets:
        headers = [_norm_header(ws.cell(1, c).value) for c in range(1, min(ws.max_column, 30) + 1)]
        header_set = {h for h in headers if h}
        if all(req in header_set for req in _REQUIRED):
            candidates.append(ws)
    if not candidates:
        raise DnkAllInventoryError(
            "Could not find an inventory sheet with columns: "
            "PONumber, Cross Reference No, Remaining Order Qty, Available Qty."
        )
    # Prefer sheets whose name looks like the master dump
    for ws in candidates:
        name = (ws.title or "").lower()
        if "availinventory" in name.replace(" ", "") or "avail" in name:
            return ws
    return candidates[0]


def _map_headers(ws: Worksheet) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for c in range(1, min(ws.max_column, 40) + 1):
        key = _norm_header(ws.cell(1, c).value)
        if key and key not in mapping:
            mapping[key] = c
    missing = [req for req in _REQUIRED if req not in mapping]
    if missing:
        raise DnkAllInventoryError(
            "Inventory sheet is missing required columns: " + ", ".join(missing)
        )
    return mapping


def _header_col(mapping: Dict[str, int], *aliases: str) -> Optional[int]:
    for alias in aliases:
        col = mapping.get(_norm_header(alias))
        if col:
            return col
    return None


def _read_rows(ws: Worksheet, mapping: Dict[str, int]) -> List[Dict[str, Any]]:
    col_customer = _header_col(mapping, "Sell to Customer No ", "Sell to Customer No")
    col_desc = _header_col(mapping, "Description")
    col_so = _header_col(mapping, "Dansko Order No")
    col_po = mapping["ponumber"]
    col_upc = mapping["cross reference no"]
    col_rem = mapping["remaining order qty"]
    col_avail = mapping["available qty"]
    col_ship = _header_col(mapping, "Req Shipment Date")
    col_item = _header_col(mapping, "Item No")
    col_variant = _header_col(mapping, "Variant Code")
    col_len = _header_col(mapping, "Length")
    col_width = _header_col(mapping, "Width")
    col_height = _header_col(mapping, "Height")
    col_weight = _header_col(mapping, "Weight")

    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        po_raw = ws.cell(r, col_po).value
        upc_raw = ws.cell(r, col_upc).value
        # Skip fully blank trailing rows
        if (po_raw is None or str(po_raw).strip() == "") and (
            upc_raw is None or str(upc_raw).strip() == ""
        ):
            continue

        po = _as_str(po_raw, strip=True)
        if not po:
            raise DnkAllInventoryError(f"Row {r}: PONumber is required.")

        upc_str = _as_str(upc_raw, strip=True)
        if not upc_str:
            raise DnkAllInventoryError(f"Row {r}: Cross Reference No is required.")

        rem = _as_int(ws.cell(r, col_rem).value, field_name=f"Row {r} Remaining Order Qty")
        avail = _as_int(ws.cell(r, col_avail).value, field_name=f"Row {r} Available Qty")
        upc_int = _as_upc_int(upc_raw, field_name=f"Row {r} Cross Reference No")

        rows.append(
            {
                "customer": _as_str(ws.cell(r, col_customer).value, strip=True) if col_customer else "",
                "description": _as_str(ws.cell(r, col_desc).value) if col_desc else "",
                "so": _as_str(ws.cell(r, col_so).value, strip=True) if col_so else "",
                "po": po,
                "upc_str": upc_str,
                "upc_int": upc_int,
                "remaining": rem,
                "available": avail,
                "ship_date": _as_str(ws.cell(r, col_ship).value, strip=True) if col_ship else "",
                "item_no": _as_str(ws.cell(r, col_item).value, strip=True) if col_item else "",
                "variant": _as_str(ws.cell(r, col_variant).value, strip=True) if col_variant else "",
                "length": _as_number(ws.cell(r, col_len).value, field_name=f"Row {r} Length")
                if col_len
                else 0,
                "width": _as_number(ws.cell(r, col_width).value, field_name=f"Row {r} Width")
                if col_width
                else 0,
                "height": _as_number(ws.cell(r, col_height).value, field_name=f"Row {r} Height")
                if col_height
                else 0,
                "weight": _as_number(ws.cell(r, col_weight).value, field_name=f"Row {r} Weight")
                if col_weight
                else 0,
            }
        )

    if not rows:
        raise DnkAllInventoryError("Inventory sheet has no data rows.")
    return rows


def _write_main_sheet(ws: Worksheet, rows: Sequence[Dict[str, Any]], source_title: str) -> None:
    ws.title = _safe_sheet_title(source_title)[:31]

    for idx, header in enumerate(MAIN_HEADERS, start=1):
        cell = ws.cell(1, idx, header)
        cell.font = _MAIN_HEADER_FONTS.get(idx - 1, _HEADER_FONT_ARIAL)
        if (idx - 1) in _CENTER_COLS:
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.alignment = Alignment(horizontal="left")

    for r_idx, row in enumerate(rows, start=2):
        values = [
            row["customer"],
            row["description"],
            row["so"],
            row["po"],
            row["upc_str"],  # string on master sheet
            row["remaining"],  # int
            row["available"],  # int
            row["ship_date"],
            row["item_no"],
            row["variant"],
            row["length"],
            row["width"],
            row["height"],
            row["weight"],
        ]
        for c_idx, value in enumerate(values, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.font = _DATA_FONT
            cell.number_format = "General"
            if (c_idx - 1) in _CENTER_COLS:
                cell.alignment = Alignment(horizontal="center")

    for col, width in _MAIN_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    last_row = len(rows) + 1
    ws.freeze_panes = "A2"
    # Match source filter span (extends past used columns)
    ws.auto_filter.ref = f"A1:Q{last_row}"
    ws.page_setup.orientation = "portrait"
    ws.sheet_format.defaultRowHeight = 15.0


def _write_po_sheet(ws: Worksheet, po: str, rows: Sequence[Dict[str, Any]]) -> None:
    for idx, header in enumerate(PO_HEADERS, start=1):
        cell = ws.cell(1, idx, header)
        cell.font = _DATA_FONT
        cell.number_format = "General"

    for r_idx, row in enumerate(rows, start=2):
        values = [
            po,
            row["upc_int"],  # int on PO sheets
            row["remaining"],
            row["available"],
        ]
        for c_idx, value in enumerate(values, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.font = _DATA_FONT
            cell.number_format = "General"

    for col, width in _PO_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width
    ws.sheet_format.defaultRowHeight = 15.0


def _output_filename(upload_filename: str) -> str:
    name = (upload_filename or "PMSH01-AvailInventory.xlsx").strip()
    if name.lower().endswith(".xlsm"):
        stem = name[:-5]
    elif name.lower().endswith(".xlsx"):
        stem = name[:-5]
    else:
        stem = name
    stem = stem.strip() or "PMSH01-AvailInventory"
    # Avoid doubling suffix if re-processed
    if stem.lower().endswith(" allinventory") or stem.lower().endswith("_allinventory"):
        return f"{stem}.xlsx"
    return f"{stem} AllInventory.xlsx"


def generate_dnk_all_inventory(
    file_bytes: bytes,
    upload_filename: str = "PMSH01-AvailInventory.xlsx",
    *,
    now: Optional[datetime] = None,
) -> DnkAllInventoryResult:
    """Build the AllInventory workbook (master + one sheet per PONumber)."""
    if not file_bytes:
        raise DnkAllInventoryError("Uploaded file is empty.")

    try:
        src = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    except Exception as exc:  # noqa: BLE001 — surface as user-facing validation
        raise DnkAllInventoryError("Could not read the Excel file. Upload a valid .xlsx workbook.") from exc

    try:
        inventory_ws = _find_inventory_sheet(src)
        mapping = _map_headers(inventory_ws)
        rows = _read_rows(inventory_ws, mapping)
        source_title = inventory_ws.title or "PMSH01-AvailInventory"
    finally:
        src.close()

    # Group by PO in first-seen order
    by_po: Dict[str, List[Dict[str, Any]]] = {}
    po_order: List[str] = []
    for row in rows:
        po = row["po"]
        if po not in by_po:
            by_po[po] = []
            po_order.append(po)
        by_po[po].append(row)

    date_stamp = _sheet_date_stamp(now)
    out = Workbook()
    main_ws = out.active
    assert main_ws is not None
    _write_main_sheet(main_ws, rows, source_title)

    used_titles = {main_ws.title.lower()}
    for po in po_order:
        title = _safe_sheet_title(f"{po} {date_stamp}")
        n = 2
        while title.lower() in used_titles:
            suffix = f" ({n})"
            title = _safe_sheet_title(f"{po} {date_stamp}")[: 31 - len(suffix)] + suffix
            n += 1
        used_titles.add(title.lower())
        po_ws = out.create_sheet(title=title)
        _write_po_sheet(po_ws, po, by_po[po])

    buf = io.BytesIO()
    out.save(buf)
    out.close()

    available_gt_zero = sum(1 for row in rows if row["available"] > 0)
    filename = _output_filename(upload_filename)

    logger.info(
        "DNK AllInventory generated: rows=%s pos=%s avail_lines=%s stamp=%s file=%s",
        len(rows),
        len(po_order),
        available_gt_zero,
        date_stamp,
        filename,
    )

    return DnkAllInventoryResult(
        file_bytes=buf.getvalue(),
        filename=filename,
        main_sheet_name=main_ws.title,
        row_count=len(rows),
        po_count=len(po_order),
        available_gt_zero=available_gt_zero,
        date_stamp=date_stamp,
        po_names=po_order,
    )
