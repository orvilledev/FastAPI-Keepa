"""Keepa Import Export tool (standalone, read-only).

Fetches the live Keepa **buy-box winner** for a set of UPCs and builds an Excel
file that matches the Keepa desktop export layout consumed by Import Mode. Only
the six columns Import Mode reads are populated, placed at their fixed Excel
positions:

    A (1)  = Imported by Code            -> UPC
    C (3)  = Title                       -> product title
    F (6)  = Buy Box: Buy Box Seller     -> "<seller name> / <seller id>"
    H (8)  = Buy Box: Current            -> buy box price (dollars)
    L (12) = ASIN                        -> product ASIN
    U (21) = URL: Amazon                 -> Amazon product link

Goal: produce a row for every UPC like the Keepa desktop export does — as few
blank cells as possible — while keeping token use low. Every UPC is fetched with
a single cheap request:

    ``stats`` + ``buybox`` with the ``offers`` parameter omitted (see
    ``KeepaClient.fetch_buybox_only``). Keepa returns title/ASIN plus the
    buy-box seller id and price directly from the product ``stats`` object, for
    roughly one token per product. (A ``code`` lookup rejects ``offers=0`` with
    HTTP 400 and rejects ``offers`` 1–19; omitting ``offers`` is the supported
    cheap path. We do NOT request the live offer list — the buy-box snapshot in
    ``stats`` is all Import Mode needs, and it matches the manual Keepa export.)

Because Keepa's per-key token buckets are small and refill slowly, a single
sweep can leave some UPCs unfetched when a worker key momentarily runs dry. Any
UPC that came back with no product at all is treated as a transient failure and
re-fetched in additional rounds, pausing between rounds so token buckets refill.
A UPC that resolves to a product but has no buy box (no seller/price in stats)
is a real Amazon state — it is left blank in those columns, exactly like the
manual export, and is not re-fetched.

The buy-box seller *name* is resolved in this order with zero extra tokens:
Keepa seller name from the response, else the cached ``seller_names`` table
(seller id -> name), else the seller id alone.

There is no MAP comparison here — this is a pure Keepa report export. This module
does not touch the scheduler, batch jobs, the import-upload pipeline, or any
shared run state; it only reads the UPC list provided by the caller and calls
Keepa. ``CSVGenerator.extract_keepa_product_data`` is used purely as a read-only
extraction helper.
"""
from __future__ import annotations

import asyncio
import logging
from io import BytesIO
from typing import Any, Awaitable, Callable, Dict, List, Optional, Union

from openpyxl import Workbook

from app.services.csv_generator import CSVGenerator
from app.services.keepa_client import MultiKeyKeepaClient

logger = logging.getLogger(__name__)

ProgressCallback = Callable[
    [int, int, str, str, int, int],
    Union[None, Awaitable[None]],
]
# Args: fetched_count, total, phase, message, enrich_total, phase_completed

# Optional predicate the caller supplies so a build can be stopped mid-flight
# (e.g. the user clicked "Stop build"). Returns True when work should halt.
CancelCallback = Callable[[], bool]


class KeepaBuildCancelled(Exception):
    """Raised internally to unwind a build once cancellation is requested."""

# Transient Keepa failures (timeouts, or a worker key momentarily out of tokens)
# return no product at all. We re-fetch those UPCs in additional rounds, pausing
# between rounds so per-key token buckets refill, instead of silently writing a
# blank row. A UPC that resolves to a product but has no buy box is NOT
# re-fetched — that is a real Amazon state and matches the manual Keepa export.
_MAX_REFETCH_ROUNDS = 5
_REFETCH_ROUND_DELAY_SECONDS = 20

# Fixed Keepa export schema expected by Import Mode (1-based Excel columns).
_COLUMN_HEADERS: Dict[int, str] = {
    1: "Imported by Code",
    3: "Title",
    6: "Buy Box: Buy Box Seller",
    8: "Buy Box: Current",
    12: "ASIN",
    21: "URL: Amazon",
}
_MAX_COLUMN = 21


def _format_seller_cell(name: str, seller_id: str) -> str:
    """Match the Keepa export form ``<seller name> / <seller id>``."""
    name = (name or "").strip()
    seller_id = (seller_id or "").strip()
    if name and seller_id:
        return f"{name} / {seller_id}"
    return name or seller_id


def _normalize_seller_id(sid: Any) -> str:
    if sid is None:
        return ""
    return str(sid).strip()


def _extract_buybox_fields(
    keepa_data: Optional[Dict[str, Any]],
    seller_name_map: Dict[str, str],
) -> Dict[str, Any]:
    """Extract the buy-box winner from a Keepa response.

    Reuses ``CSVGenerator.extract_keepa_product_data`` for the same buy-box
    resolution the daily runs use: buy-box price/seller come from the product
    ``stats`` object (no offer list needed). The seller display name is then
    resolved Keepa-name -> cached ``seller_names`` map -> seller id, so a cached
    name fills in even when Keepa omits it.
    """
    if not keepa_data or not isinstance(keepa_data, dict):
        return {}

    fields = CSVGenerator.extract_keepa_product_data(keepa_data) or {}
    if not fields:
        return {}

    asin = str(fields.get("asin") or "").strip()
    title = str(fields.get("title") or "").strip()
    seller_id = _normalize_seller_id(fields.get("buy_box_seller_id"))

    seller_name = (fields.get("buy_box_seller_name") or "").strip()
    if seller_name.lower() == "unknown":
        seller_name = ""
    if not seller_name and seller_id:
        seller_name = seller_name_map.get(seller_id, "")

    price = fields.get("buy_box_price")
    try:
        price = float(price) if price is not None else None
        if price is not None and price <= 0:
            price = None
    except (TypeError, ValueError):
        price = None

    return {
        "asin": asin,
        "title": title,
        "buy_box_seller_id": seller_id,
        "buy_box_seller_name": seller_name,
        "buy_box_price": price,
    }


def _build_row_values(upc: str, fields: Dict[str, Any]) -> Dict[int, str]:
    """Map extracted buy-box fields to fixed Import Mode column positions."""
    asin = str(fields.get("asin") or "").strip()
    title = str(fields.get("title") or "").strip()
    seller_cell = _format_seller_cell(
        fields.get("buy_box_seller_name") or "",
        fields.get("buy_box_seller_id") or "",
    )

    price = fields.get("buy_box_price")
    try:
        price_cell = f"{float(price):.2f}" if price is not None else ""
    except (TypeError, ValueError):
        price_cell = ""

    link = f"https://www.amazon.com/dp/{asin}?psc=1" if asin else ""

    return {
        1: str(upc),
        3: title,
        6: seller_cell,
        8: price_cell,
        12: asin,
        21: link,
    }


async def _emit_progress(
    callback: Optional[ProgressCallback],
    fetched_count: int,
    total: int,
    phase: str,
    message: str,
    enrich_total: int,
    phase_completed: int,
) -> None:
    if not callback:
        return
    result = callback(
        fetched_count, total, phase, message, enrich_total, phase_completed
    )
    if asyncio.iscoroutine(result):
        await result


def _has_product(fields: Optional[Dict[str, Any]]) -> bool:
    """True when Keepa actually returned a product (title or ASIN present).

    A UPC with neither is a transient fetch failure worth re-fetching. A UPC
    that has a product but no buy box is a real state and is left as-is.
    """
    if not fields:
        return False
    return bool(fields.get("title") or fields.get("asin"))


def _count_fetched_products(
    results: Dict[str, Dict[str, Any]],
    all_upcs: List[str],
) -> int:
    """Count UPCs that already have title or ASIN from Keepa."""
    return sum(1 for upc in all_upcs if _has_product(results.get(upc)))


def _build_multi_client() -> MultiKeyKeepaClient:
    """Create the multi-key client for Import File fetches.

    Uses the dedicated ``KEEPA_IMPORT_API_KEYS`` pool when configured (e.g. the
    few high-refill keys) so large vendor builds finish in one pass, and falls
    back to the full Keepa key pool otherwise.
    """
    import_keys = MultiKeyKeepaClient.load_import_api_keys()
    if import_keys:
        return MultiKeyKeepaClient(api_keys=import_keys)
    return MultiKeyKeepaClient()


async def _run_buybox_pass(
    upcs: List[str],
    name_map: Dict[str, str],
    results: Dict[str, Dict[str, Any]],
    *,
    all_upcs: List[str],
    total_upcs: int,
    phase: str,
    label: str,
    on_progress: Optional[ProgressCallback] = None,
    should_cancel: Optional[CancelCallback] = None,
) -> None:
    """Fetch each UPC with the cheap buy-box-only request and merge the result.

    Never overwrites an existing product row with an empty refetch result, so a
    transient failure in a later round cannot wipe out data already obtained.
    """
    if not upcs:
        return

    pass_total = len(upcs)
    progress_lock = asyncio.Lock()
    completed = 0

    multi_client = _build_multi_client()
    items = [{"upc": u} for u in upcs]

    async def process_fn(keepa_client, item) -> bool:
        nonlocal completed
        # Stop touching Keepa as soon as cancellation is requested; remaining
        # queued UPCs drain instantly without spending tokens.
        if should_cancel is not None and should_cancel():
            return False
        upc = str(item.get("upc") or "").strip()
        if not upc:
            return False
        try:
            keepa_data = await keepa_client.fetch_buybox_only(upc)
        except Exception as exc:  # defensive: never fail the whole file for one UPC
            logger.warning("Keepa buy-box fetch failed for UPC %s: %s", upc, exc)
            keepa_data = None

        fields = _extract_buybox_fields(keepa_data, name_map)
        existing = results.get(upc)
        if existing is None or _has_product(fields):
            results[upc] = fields

        async with progress_lock:
            completed += 1
            current = completed
        fetched_count = _count_fetched_products(results, all_upcs)
        await _emit_progress(
            on_progress,
            fetched_count,
            total_upcs,
            phase,
            f"{label} {current}/{pass_total}",
            enrich_total=pass_total,
            phase_completed=current,
        )
        return True

    # offers_limit=0 only tunes request pacing here; fetch_buybox_only omits the
    # offers parameter from the actual Keepa request.
    await multi_client.process_items_parallel(
        items=items, process_fn=process_fn, offers_limit=0
    )


async def _fetch_fields_for_upcs(
    upcs: List[str],
    seller_name_map: Optional[Dict[str, str]] = None,
    on_progress: Optional[ProgressCallback] = None,
    should_cancel: Optional[CancelCallback] = None,
) -> Dict[str, Dict[str, Any]]:
    """Fetch buy-box data for every UPC, retrying transient failures.

    One cheap buy-box-only sweep over all UPCs, then up to ``_MAX_REFETCH_ROUNDS``
    additional rounds that re-fetch only the UPCs that returned no product,
    pausing between rounds so per-key token buckets refill. Creates its own
    multi-key client per round so it never alters shared run state; workers are
    asyncio tasks on a single thread, so dict writes are safe.
    """
    results: Dict[str, Dict[str, Any]] = {}
    name_map = seller_name_map or {}
    total_upcs = len(upcs)

    def _cancelled() -> bool:
        return should_cancel is not None and should_cancel()

    await _run_buybox_pass(
        upcs,
        name_map,
        results,
        all_upcs=upcs,
        total_upcs=total_upcs,
        phase="pass1",
        label="Fetching",
        on_progress=on_progress,
        should_cancel=should_cancel,
    )

    if _cancelled():
        raise KeepaBuildCancelled()

    for round_num in range(1, _MAX_REFETCH_ROUNDS + 1):
        missing = [u for u in upcs if not _has_product(results.get(u))]
        if not missing:
            break
        logger.info(
            "Keepa Import File: re-fetch round %d for %d UPC(s) with no product",
            round_num,
            len(missing),
        )
        await _emit_progress(
            on_progress,
            _count_fetched_products(results, upcs),
            total_upcs,
            "pass2",
            f"Retrying {len(missing)} UPC(s) (round {round_num})…",
            enrich_total=len(missing),
            phase_completed=0,
        )
        # Let per-key token buckets refill before hammering the same keys again,
        # but wake early in short slices so a cancel is honored quickly.
        for _ in range(_REFETCH_ROUND_DELAY_SECONDS):
            if _cancelled():
                raise KeepaBuildCancelled()
            await asyncio.sleep(1)
        await _run_buybox_pass(
            missing,
            name_map,
            results,
            all_upcs=upcs,
            total_upcs=total_upcs,
            phase="pass2",
            label=f"Retry round {round_num}:",
            on_progress=on_progress,
            should_cancel=should_cancel,
        )
        if _cancelled():
            raise KeepaBuildCancelled()

    return results


def build_workbook_bytes(
    upcs: List[str],
    fields_by_upc: Dict[str, Dict[str, Any]],
    include_header: bool = True,
) -> bytes:
    """Render the Keepa-format Excel workbook to bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Keepa Export"

    if include_header:
        for col in range(1, _MAX_COLUMN + 1):
            ws.cell(row=1, column=col, value=_COLUMN_HEADERS.get(col, ""))

    start_row = 2 if include_header else 1
    for offset, upc in enumerate(upcs):
        row_idx = start_row + offset
        values = _build_row_values(upc, fields_by_upc.get(upc, {}))
        for col, val in values.items():
            ws.cell(row=row_idx, column=col, value=val)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


async def generate_keepa_import_file(
    upcs: List[str],
    seller_name_map: Optional[Dict[str, str]] = None,
    include_header: bool = True,
    on_progress: Optional[ProgressCallback] = None,
    should_cancel: Optional[CancelCallback] = None,
) -> bytes:
    """Fetch buy-box data for ``upcs`` and return an Import Mode-ready .xlsx file.

    ``seller_name_map`` (seller id -> name) fills in the buy-box seller display
    name without spending extra Keepa tokens. Every UPC is fetched with the cheap
    buy-box-only request; UPCs that return no product are retried in additional
    rounds before being left blank.
    """
    fields_by_upc = await _fetch_fields_for_upcs(
        upcs,
        seller_name_map,
        on_progress=on_progress,
        should_cancel=should_cancel,
    )
    fetched = _count_fetched_products(fields_by_upc, upcs)
    await _emit_progress(
        on_progress,
        fetched,
        len(upcs),
        "excel",
        "Building Excel file…",
        enrich_total=0,
        phase_completed=1,
    )
    return build_workbook_bytes(upcs, fields_by_upc, include_header=include_header)
