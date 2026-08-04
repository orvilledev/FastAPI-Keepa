"""Build Master Sheet output from uploaded rows + UPC/DIMS catalogs.

Mirrors the Excel template formulas on MASTER SHEET:

- S/C/S = STYLE & \" \" & COLOR & \" \" & SIZE
- UPC   = VLOOKUP(S/C/S, UPC catalog S/C/S → External ID / UPC Code)
- MC L/W/H = VLOOKUP(UPC, DIMS by UPC #), else match Description + size
  with M/W/K/T (and KIDS') gender prefix sensitivity.
"""
from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from supabase import Client

logger = logging.getLogger(__name__)

MASTER_SHEET_NAME = "MASTER SHEET"
MASTER_HEADERS = [
    "WEIGHT (lbs)",
    "MC Length",
    "MC Width",
    "MC Height",
    "BOX#",
    "WEIGHT (lbs)",
    "PO# / ORD#",
    "CARTON# / TICKET#",
    "STYLE",
    "COLOR",
    "DESCRIPTION",
    "SIZE",
    "S/C/S",
    "UPC",
    "TOT QTY",
    "RECEIVE AT FC",
]

# Input column aliases (case-insensitive) → canonical field
_INPUT_ALIASES = {
    "box#": "box",
    "box": "box",
    "weight (lbs)": "weight",
    "weight": "weight",
    "po# / ord#": "po",
    "po#": "po",
    "po": "po",
    "ord#": "po",
    "carton# / ticket#": "carton",
    "carton#": "carton",
    "ticket#": "carton",
    "carton": "carton",
    "style": "style",
    "color": "color",
    "description": "description",
    "size": "size",
    "tot qty": "qty",
    "qty": "qty",
    "quantity": "qty",
    "receive at fc": "receive_at_fc",
    "s/c/s": "scs",
    "upc": "upc",
}

_GENDER_PREFIXES = ("M", "W", "K", "T")


@dataclass
class MasterSheetResult:
    file_bytes: bytes
    filename: str
    total_rows: int
    upc_matched: int
    upc_missing: int
    mc_matched_by_upc: int
    mc_matched_by_desc_size: int
    mc_missing: int
    warnings: List[str] = field(default_factory=list)


def _cell_str(raw: Any) -> str:
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return "TRUE" if raw else "FALSE"
    if isinstance(raw, int):
        return str(raw)
    if isinstance(raw, float):
        if raw.is_integer():
            return str(int(raw))
        return format(raw, ".15g").strip()
    text = str(raw).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        return text[:-2]
    return text


def _normalize_header(cell: Any) -> str:
    return re.sub(r"\s+", " ", _cell_str(cell)).strip().lower()


def size_code_from_master_size(size: str) -> str:
    """Map Master SIZE like '8 M' / '7.5 M' to DIMS SKU suffix '08' / '07.5'."""
    s = _cell_str(size).upper()
    if not s:
        return ""
    tok = s.split()[0]
    if tok.replace(".", "", 1).isdigit():
        if "." in tok:
            whole, frac = tok.split(".", 1)
            return f"{int(whole):02d}.{frac}"
        return f"{int(tok):02d}"
    return tok


def gender_prefix(description: str) -> str:
    """Return M/W/K/T when description starts with those (KIDS' counts as K)."""
    u = _cell_str(description).upper()
    if not u:
        return ""
    if u.startswith("KIDS'") or u.startswith("KIDS "):
        return "K"
    first = u.split(" ", 1)[0]
    if first in _GENDER_PREFIXES:
        return first
    return ""


def description_match_key(description: str) -> str:
    """Exact description match key; preserves M/W/K/T (normalizes KIDS' → K)."""
    text = _cell_str(description)
    if not text:
        return ""
    upper = text.upper()
    if upper.startswith("KIDS'"):
        rest = text[5:].lstrip()
        return f"K {rest}".upper()
    if upper.startswith("KIDS "):
        rest = text[5:].lstrip()
        return f"K {rest}".upper()
    return upper


def description_match_keys(description: str) -> List[str]:
    """Primary description key plus aliases (e.g. Tasman Nubuck → Tasman II)."""
    primary = description_match_key(description)
    if not primary:
        return []
    keys = [primary]
    # "W TASMAN NUBUCK" / "M TASMAN II NUBUCK" → same gender + "TASMAN II"
    if "TASMAN" in primary and "NUBUCK" in primary:
        prefix = gender_prefix(primary)
        alias = f"{prefix} TASMAN II".strip() if prefix else "TASMAN II"
        if alias not in keys:
            keys.append(alias)
    return keys


def build_scs(style: str, color: str, size: str) -> str:
    """Excel: =(STYLE&\" \"&COLOR)&\" \"&SIZE"""
    return f"{_cell_str(style)} {_cell_str(color)} {_cell_str(size)}".strip()


def _fetch_all(db: Client, table: str, columns: str) -> List[dict]:
    """Page through PostgREST results (supabase-py range end is exclusive here)."""
    rows: List[dict] = []
    page_size = 1000
    offset = 0
    while True:
        chunk = (
            db.table(table)
            .select(columns)
            .range(offset, offset + page_size)
            .execute()
            .data
            or []
        )
        if not chunk:
            break
        rows.extend(chunk)
        offset += len(chunk)
        if len(chunk) < page_size:
            break
    return rows


def load_upc_index(db: Client) -> Dict[str, str]:
    """S/C/S → UPC code."""
    rows = _fetch_all(db, "catalog_upc_records", "scs,upc_code")
    index: Dict[str, str] = {}
    for row in rows:
        scs = _cell_str(row.get("scs"))
        upc = _cell_str(row.get("upc_code"))
        if scs and upc:
            index[scs] = upc
    return index


def load_dims_indexes(
    db: Client,
) -> Tuple[Dict[str, Tuple[str, str, str]], Dict[Tuple[str, str], Tuple[str, str, str]]]:
    """Return (by_upc, by_desc_size) → (mc_length, mc_width, mc_height)."""
    rows = _fetch_all(
        db, "catalog_dims_records", "upc_number,sku,description,row_data"
    )
    by_upc: Dict[str, Tuple[str, str, str]] = {}
    by_desc_size: Dict[Tuple[str, str], Tuple[str, str, str]] = {}
    for row in rows:
        rd = row.get("row_data") or {}
        mc_l = _cell_str(rd.get("MC Length"))
        mc_w = _cell_str(rd.get("MC Width"))
        mc_h = _cell_str(rd.get("MC Height"))
        if not (mc_l and mc_w and mc_h):
            continue
        mc = (mc_l, mc_w, mc_h)
        upc = _cell_str(row.get("upc_number") or rd.get("UPC #"))
        if upc:
            by_upc[upc] = mc
        desc_key = description_match_key(row.get("description") or rd.get("Description"))
        sku = _cell_str(row.get("sku") or rd.get("SKU"))
        size_suffix = sku.rsplit("-", 1)[-1].strip() if "-" in sku else ""
        if desc_key and size_suffix:
            by_desc_size.setdefault((desc_key, size_suffix), mc)
    return by_upc, by_desc_size


def _find_master_sheet(wb):
    for name in wb.sheetnames:
        if name.strip().upper() == MASTER_SHEET_NAME:
            return wb[name]
    return wb[wb.sheetnames[0]]


def _header_map(header_row: Tuple[Any, ...]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    weight_idxs: List[int] = []
    for idx, cell in enumerate(header_row):
        key = _INPUT_ALIASES.get(_normalize_header(cell))
        if not key:
            continue
        if key == "weight":
            weight_idxs.append(idx)
            continue
        if key not in mapping:
            mapping[key] = idx
    # Prefer the second WEIGHT column (column F in the template) as carton weight input.
    if weight_idxs:
        mapping["weight"] = weight_idxs[-1] if len(weight_idxs) > 1 else weight_idxs[0]
    return mapping


def _row_values(row: Tuple[Any, ...], mapping: Dict[str, int]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for key, idx in mapping.items():
        out[key] = _cell_str(row[idx]) if idx < len(row) else ""
    return out


def parse_master_input(filename: str, content: bytes) -> List[Dict[str, str]]:
    lower = (filename or "").lower()
    if not lower.endswith((".xlsx", ".xlsm", ".xls")):
        raise ValueError("Upload an .xlsx Master Sheet file.")

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = _find_master_sheet(wb)
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return []
        mapping = _header_map(tuple(header))
        required = ("style", "color", "description", "size")
        missing = [k for k in required if k not in mapping]
        if missing:
            raise ValueError(
                "Master Sheet must include STYLE, COLOR, DESCRIPTION, and SIZE columns "
                f"(missing: {', '.join(missing)})."
            )

        parsed: List[Dict[str, str]] = []
        for row in rows_iter:
            values = _row_values(tuple(row), mapping)
            if not any(values.get(k) for k in ("style", "color", "description", "size", "box")):
                continue
            if not values.get("style"):
                continue
            parsed.append(values)
        return parsed
    finally:
        wb.close()


def _lookup_mc(
    upc: str,
    description: str,
    size: str,
    by_upc: Dict[str, Tuple[str, str, str]],
    by_desc_size: Dict[Tuple[str, str], Tuple[str, str, str]],
) -> Tuple[Optional[Tuple[str, str, str]], str]:
    if upc and upc in by_upc:
        return by_upc[upc], "upc"
    sc = size_code_from_master_size(size)
    size_candidates = [sc]
    tok = _cell_str(size).split()[0] if _cell_str(size) else ""
    if tok and tok not in size_candidates:
        size_candidates.append(tok)
    for desc_key in description_match_keys(description):
        if not desc_key:
            continue
        for size_key in size_candidates:
            if not size_key:
                continue
            hit = by_desc_size.get((desc_key, size_key))
            if hit:
                return hit, "desc_size"
    return None, ""


def build_master_sheet_workbook(
    input_rows: List[Dict[str, str]],
    upc_by_scs: Dict[str, str],
    dims_by_upc: Dict[str, Tuple[str, str, str]],
    dims_by_desc_size: Dict[Tuple[str, str], Tuple[str, str, str]],
) -> MasterSheetResult:
    wb = Workbook()
    ws = wb.active
    ws.title = MASTER_SHEET_NAME
    ws.append(MASTER_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    upc_matched = 0
    upc_missing = 0
    mc_by_upc = 0
    mc_by_desc = 0
    mc_missing = 0
    warnings: List[str] = []

    for idx, row in enumerate(input_rows, start=1):
        style = row.get("style", "")
        color = row.get("color", "")
        description = row.get("description", "")
        size = row.get("size", "")
        weight = row.get("weight", "")
        box = row.get("box", "") or str(idx)
        po = row.get("po", "")
        carton = row.get("carton", "")
        qty = row.get("qty", "")
        receive = row.get("receive_at_fc", "")

        scs = build_scs(style, color, size)
        upc = upc_by_scs.get(scs, "")
        if not upc and row.get("upc"):
            # Allow pre-filled UPC on the upload, still validate against blanks.
            upc = row["upc"]
        if upc:
            upc_matched += 1
        else:
            upc_missing += 1
            warnings.append(f"Row {idx}: missing UPC for S/C/S '{scs}'")

        mc, how = _lookup_mc(upc, description, size, dims_by_upc, dims_by_desc_size)
        if mc:
            if how == "upc":
                mc_by_upc += 1
            else:
                mc_by_desc += 1
            mc_l, mc_w, mc_h = mc
        else:
            mc_missing += 1
            mc_l = mc_w = mc_h = ""
            if upc:
                warnings.append(
                    f"Row {idx}: no MC dims for UPC {upc} "
                    f"(desc '{description}', size '{size}', gender {gender_prefix(description) or 'n/a'})"
                )

        # Never write N/A — blank when unmatched.
        ws.append(
            [
                weight,  # A mirrors F
                mc_l,
                mc_w,
                mc_h,
                box,
                weight,
                po,
                carton,
                style,
                color,
                description,
                size,
                scs,
                upc,
                qty,
                receive,
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return MasterSheetResult(
        file_bytes=buf.getvalue(),
        filename="Master_Sheet.xlsx",
        total_rows=len(input_rows),
        upc_matched=upc_matched,
        upc_missing=upc_missing,
        mc_matched_by_upc=mc_by_upc,
        mc_matched_by_desc_size=mc_by_desc,
        mc_missing=mc_missing,
        warnings=warnings[:50],
    )


def generate_master_sheet(db: Client, filename: str, content: bytes) -> MasterSheetResult:
    input_rows = parse_master_input(filename, content)
    if not input_rows:
        raise ValueError("No Master Sheet data rows found in the uploaded file.")

    upc_by_scs = load_upc_index(db)
    if not upc_by_scs:
        raise ValueError(
            "UPC catalog is empty. Upload the UPC catalog in the UPC sidebar first."
        )
    dims_by_upc, dims_by_desc_size = load_dims_indexes(db)
    if not dims_by_upc and not dims_by_desc_size:
        raise ValueError(
            "DIMS catalog is empty. Upload the DIMS catalog in the DIMS sidebar first."
        )

    return build_master_sheet_workbook(
        input_rows, upc_by_scs, dims_by_upc, dims_by_desc_size
    )


def build_master_sheet_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = MASTER_SHEET_NAME
    ws.append(MASTER_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
