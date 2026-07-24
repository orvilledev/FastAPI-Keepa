"""Generate Amazon Send-to-Amazon FBA manifest workbooks from a packing sheet.

Input columns (header row, case-insensitive):
  UPC Code | Vendor | Employee | Qty Per Box | Total QTY | Pack Group # | Amazon Link

Rules:
  - Pack Group # is forward-filled down contiguous rows.
  - All vendors are merged by pack group number.
  - Duplicate UPCs within a pack group are aggregated (first-seen order, summed qty).
  - One STA workbook is emitted per pack group.
  - Zip / file prefix uses the first vendor with a trailing \" PLACE\" stripped.
  - Date stamp uses M.D.YY with no zero-padding (e.g. 7.24.26).
"""
from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import BinaryIO, Sequence

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

BACKEND_ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_PATH = BACKEND_ROOT / "assets" / "amazon_sta_manifest_template.xlsx"

# Amazon STA sheet tabs use an en-dash (U+2013), not a hyphen.
TEMPLATE_SHEET_NAME = "Create workflow \u2013 template"

REQUIRED_HEADERS = {
    "upc": {"upc code", "upc", "merchant sku", "sku", "fnsku"},
    "vendor": {"vendor"},
    "total_qty": {"total qty", "total quantity", "quantity", "qty", "total qyt"},
    "pack_group": {"pack group #", "pack group", "pack group number", "pg", "packgroup"},
}

OPTIONAL_HEADERS = {
    "employee": {"employee"},
    "qty_per_box": {"qty per box", "qty  per box", "units per box"},
    "amazon_link": {"amazon link"},
}

_MAX_UPLOAD_BYTES = 15 * 1024 * 1024
_SKU_FONT = Font(name="Calibri", size=11)
_SKU_ALIGN = Alignment(horizontal="center", vertical="center")
_QTY_FONT = Font(name="Calibri", size=11)
_QTY_ALIGN = Alignment(horizontal="center", vertical="center")


class ManifestGeneratorError(ValueError):
    """Raised for user-correctable input problems."""


@dataclass(frozen=True)
class ManifestRow:
    upc: str
    vendor: str
    total_qty: int
    pack_group: int | None


@dataclass(frozen=True)
class ManifestBuildResult:
    zip_bytes: bytes
    zip_filename: str
    file_count: int
    pack_groups: list[int]
    primary_vendor: str
    sku_count: int
    total_units: int


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def _find_column_map(headers: Sequence[object]) -> dict[str, int]:
    normalized = [_normalize_header(h) for h in headers]
    col_map: dict[str, int] = {}

    for key, aliases in {**REQUIRED_HEADERS, **OPTIONAL_HEADERS}.items():
        for idx, header in enumerate(normalized):
            if header in aliases:
                col_map[key] = idx
                break

    missing = [k for k in ("upc", "vendor", "total_qty", "pack_group") if k not in col_map]
    if missing:
        pretty = {
            "upc": "UPC Code",
            "vendor": "Vendor",
            "total_qty": "Total QTY",
            "pack_group": "Pack Group #",
        }
        labels = ", ".join(pretty[m] for m in missing)
        raise ManifestGeneratorError(
            f"Missing required column(s): {labels}. "
            "Expected headers like UPC Code, Vendor, Total QTY, Pack Group #."
        )
    return col_map


def _parse_qty(value: object, row_number: int) -> int:
    if value is None or value == "":
        raise ManifestGeneratorError(f"Row {row_number}: Total QTY is required.")
    if isinstance(value, bool):
        raise ManifestGeneratorError(f"Row {row_number}: Total QTY must be a whole number.")
    if isinstance(value, int):
        qty = value
    elif isinstance(value, float):
        if not value.is_integer():
            raise ManifestGeneratorError(f"Row {row_number}: Total QTY must be a whole number.")
        qty = int(value)
    else:
        text = str(value).strip().replace(",", "")
        if not re.fullmatch(r"-?\d+", text):
            raise ManifestGeneratorError(f"Row {row_number}: Total QTY must be a whole number.")
        qty = int(text)
    if qty < 1:
        raise ManifestGeneratorError(f"Row {row_number}: Total QTY must be 1 or greater.")
    return qty


def _parse_pack_group(value: object, row_number: int) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        raise ManifestGeneratorError(f"Row {row_number}: Pack Group # must be a whole number.")
    if isinstance(value, int):
        pg = value
    elif isinstance(value, float):
        if not value.is_integer():
            raise ManifestGeneratorError(f"Row {row_number}: Pack Group # must be a whole number.")
        pg = int(value)
    else:
        text = str(value).strip()
        if not re.fullmatch(r"-?\d+", text):
            raise ManifestGeneratorError(f"Row {row_number}: Pack Group # must be a whole number.")
        pg = int(text)
    if pg < 1:
        raise ManifestGeneratorError(f"Row {row_number}: Pack Group # must be 1 or greater.")
    return pg


def normalize_vendor_label(vendor: str) -> str:
    """Strip a trailing ' PLACE' suffix used for warehouse-placement rows."""
    text = (vendor or "").strip()
    if re.search(r"\s+PLACE$", text, flags=re.IGNORECASE):
        text = re.sub(r"\s+PLACE$", "", text, flags=re.IGNORECASE).strip()
    return text or "MANIFEST"


def format_manifest_date(when: date | datetime | None = None) -> str:
    """Return M.D.YY with no zero-padding (matches sample: 7.24.26)."""
    d = when.date() if isinstance(when, datetime) else (when or date.today())
    return f"{d.month}.{d.day}.{d.year % 100}"


def parse_manifest_rows(file_obj: BinaryIO | bytes) -> list[ManifestRow]:
    raw = file_obj if isinstance(file_obj, (bytes, bytearray)) else file_obj.read()
    if not raw:
        raise ManifestGeneratorError("Uploaded file is empty.")
    if len(raw) > _MAX_UPLOAD_BYTES:
        raise ManifestGeneratorError("File is too large (max 15 MB).")

    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 - surface as user error
        raise ManifestGeneratorError("Could not read the Excel file. Upload a valid .xlsx workbook.") from exc

    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise ManifestGeneratorError("The workbook has no header row.") from exc

        col_map = _find_column_map(header_row)
        parsed: list[ManifestRow] = []
        for offset, values in enumerate(rows_iter, start=2):
            if values is None:
                continue
            upc_raw = values[col_map["upc"]] if col_map["upc"] < len(values) else None
            if upc_raw is None or str(upc_raw).strip() == "":
                # Skip fully empty trailing rows; error if other columns have data.
                vendor_raw = values[col_map["vendor"]] if col_map["vendor"] < len(values) else None
                qty_raw = values[col_map["total_qty"]] if col_map["total_qty"] < len(values) else None
                if (vendor_raw is None or str(vendor_raw).strip() == "") and (
                    qty_raw is None or qty_raw == ""
                ):
                    continue
                raise ManifestGeneratorError(f"Row {offset}: UPC Code is required.")

            upc = str(upc_raw).strip()
            # Avoid scientific notation artifacts from numeric Excel cells.
            if isinstance(upc_raw, float) and upc_raw.is_integer():
                upc = str(int(upc_raw))
            elif isinstance(upc_raw, int):
                upc = str(upc_raw)

            vendor_val = values[col_map["vendor"]] if col_map["vendor"] < len(values) else None
            vendor = str(vendor_val or "").strip()
            if not vendor:
                raise ManifestGeneratorError(f"Row {offset}: Vendor is required.")

            qty = _parse_qty(
                values[col_map["total_qty"]] if col_map["total_qty"] < len(values) else None,
                offset,
            )
            pack_group = _parse_pack_group(
                values[col_map["pack_group"]] if col_map["pack_group"] < len(values) else None,
                offset,
            )
            parsed.append(ManifestRow(upc=upc, vendor=vendor, total_qty=qty, pack_group=pack_group))
    finally:
        wb.close()

    if not parsed:
        raise ManifestGeneratorError("No data rows found in the workbook.")
    return parsed


def forward_fill_pack_groups(rows: Sequence[ManifestRow]) -> list[tuple[ManifestRow, int]]:
    filled: list[tuple[ManifestRow, int]] = []
    current: int | None = None
    for idx, row in enumerate(rows, start=2):
        if row.pack_group is not None:
            current = row.pack_group
        if current is None:
            raise ManifestGeneratorError(
                f"Row {idx}: Pack Group # is missing. Provide a Pack Group # on the first data row."
            )
        filled.append((row, current))
    return filled


def aggregate_by_pack_group(
    rows: Sequence[ManifestRow],
) -> tuple[str, dict[int, list[tuple[str, int]]]]:
    filled = forward_fill_pack_groups(rows)
    primary_vendor = normalize_vendor_label(filled[0][0].vendor)

    # pack_group -> ordered upc list + quantities
    order: dict[int, list[str]] = {}
    qty: dict[int, dict[str, int]] = {}
    for row, pg in filled:
        if pg not in order:
            order[pg] = []
            qty[pg] = {}
        if row.upc not in qty[pg]:
            order[pg].append(row.upc)
            qty[pg][row.upc] = 0
        qty[pg][row.upc] += row.total_qty

    grouped = {pg: [(upc, qty[pg][upc]) for upc in order[pg]] for pg in sorted(order.keys())}
    return primary_vendor, grouped


def _template_sheet(wb: Workbook) -> Worksheet:
    if TEMPLATE_SHEET_NAME in wb.sheetnames:
        return wb[TEMPLATE_SHEET_NAME]
    for name in wb.sheetnames:
        if "template" in name.lower():
            return wb[name]
    raise ManifestGeneratorError("Amazon STA template is missing the Create workflow template sheet.")


def _write_sku_rows(ws: Worksheet, items: Sequence[tuple[str, int]]) -> None:
    start_row = 9
    for offset, (sku, quantity) in enumerate(items):
        row_idx = start_row + offset
        sku_cell = ws.cell(row=row_idx, column=1, value=sku)
        sku_cell.font = _SKU_FONT
        sku_cell.alignment = _SKU_ALIGN
        sku_cell.number_format = numbers.FORMAT_TEXT

        qty_cell = ws.cell(row=row_idx, column=2, value=int(quantity))
        qty_cell.font = _QTY_FONT
        qty_cell.alignment = _QTY_ALIGN
        qty_cell.number_format = "General"


def build_pack_group_workbook(items: Sequence[tuple[str, int]], template_path: Path | None = None) -> bytes:
    path = template_path or TEMPLATE_PATH
    if not path.exists():
        raise ManifestGeneratorError("Manifest template asset is missing on the server.")

    wb = load_workbook(path)
    try:
        ws = _template_sheet(wb)
        _write_sku_rows(ws, items)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    finally:
        wb.close()


def build_manifest_zip(
    file_obj: BinaryIO | bytes,
    *,
    when: date | datetime | None = None,
    template_path: Path | None = None,
) -> ManifestBuildResult:
    rows = parse_manifest_rows(file_obj)
    primary_vendor, grouped = aggregate_by_pack_group(rows)
    date_stamp = format_manifest_date(when)

    zip_buf = io.BytesIO()
    sku_count = 0
    total_units = 0
    pack_groups = list(grouped.keys())

    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for pg, items in grouped.items():
            workbook_bytes = build_pack_group_workbook(items, template_path=template_path)
            filename = f"{primary_vendor} PG{pg} {date_stamp}.xlsx"
            zf.writestr(filename, workbook_bytes)
            sku_count += len(items)
            total_units += sum(q for _, q in items)

    zip_filename = f"{primary_vendor} FBA Manifests {date_stamp}.zip"
    return ManifestBuildResult(
        zip_bytes=zip_buf.getvalue(),
        zip_filename=zip_filename,
        file_count=len(pack_groups),
        pack_groups=pack_groups,
        primary_vendor=primary_vendor,
        sku_count=sku_count,
        total_units=total_units,
    )
