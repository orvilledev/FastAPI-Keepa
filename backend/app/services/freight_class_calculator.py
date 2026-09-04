"""NMFC density-based freight class calculator (XPO-compatible logic)."""
from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

CUBIC_INCHES_PER_CU_FT = 1728

# NMFC 13-sub density scale (effective July 2025): lower bound inclusive.
_DENSITY_BANDS: list[tuple[float, float]] = [
    (50.0, 50),
    (35.0, 55),
    (30.0, 60),
    (22.5, 65),
    (15.0, 70),
    (12.0, 85),
    (10.0, 92.5),
    (8.0, 100),
    (6.0, 125),
    (4.0, 175),
    (2.0, 250),
    (1.0, 300),
    (0.0, 400),
]

_HEADER_ALIASES: dict[str, str] = {
    "shipment id": "shipment_id",
    "shipment_id": "shipment_id",
    "shipment": "shipment_id",
    # Pallet Count = how many identical pallets this row represents (volume multiplier).
    # Pallet Number (#1, #2, …) is a line label only — never aliased here.
    "pallet count": "pallets",
    "pallet_count": "pallets",
    "pallets": "pallets",
    "pieces": "pallets",
    "piece count": "pallets",
    "piece_cnt": "pallets",
    "weight": "weight",
    "weight (lbs)": "weight",
    "weight lbs": "weight",
    "length": "length",
    "length (in)": "length",
    "width": "width",
    "width (in)": "width",
    "height": "height",
    "height (in)": "height",
}


class FreightClassCalculatorError(Exception):
    """Domain error for freight class calculator inputs."""


@dataclass
class ShipmentLineItem:
    pallets: int
    weight_lbs: float
    length_in: float
    width_in: float
    height_in: float
    adjusted_height_in: float | None = None
    height_rule_applied: bool = False
    cubic_feet: float = 0.0

    def to_dict(self) -> dict[str, Any]:
        return {
            "pallets": self.pallets,
            "weight_lbs": round(self.weight_lbs, 2),
            "length_in": round(self.length_in, 2),
            "width_in": round(self.width_in, 2),
            "height_in": round(self.height_in, 2),
            "adjusted_height_in": (
                round(self.adjusted_height_in, 2) if self.adjusted_height_in is not None else None
            ),
            "height_rule_applied": self.height_rule_applied,
            "cubic_feet": round(self.cubic_feet, 4),
        }


@dataclass
class ShipmentResult:
    shipment_id: str
    line_items: list[ShipmentLineItem] = field(default_factory=list)
    total_weight_lbs: float = 0.0
    total_cubic_feet: float = 0.0
    density_pcf: float = 0.0
    freight_class: float = 400
    height_rule_applied: bool = False

    def to_dict(self) -> dict[str, Any]:
        return {
            "shipment_id": self.shipment_id,
            "line_items": [item.to_dict() for item in self.line_items],
            "total_weight_lbs": round(self.total_weight_lbs, 2),
            "total_cubic_feet": round(self.total_cubic_feet, 4),
            "density_pcf": round(self.density_pcf, 4),
            "freight_class": self.freight_class,
            "height_rule_applied": self.height_rule_applied,
        }


@dataclass
class CalculationResult:
    shipments: list[ShipmentResult]

    def to_dict(self) -> dict[str, Any]:
        class_breakdown: dict[str, int] = {}
        for shipment in self.shipments:
            key = str(shipment.freight_class)
            class_breakdown[key] = class_breakdown.get(key, 0) + 1
        return {
            "shipments": [s.to_dict() for s in self.shipments],
            "summary": {
                "shipment_count": len(self.shipments),
                "class_breakdown": class_breakdown,
            },
        }


def density_to_freight_class(pcf: float) -> float:
    """Map pounds-per-cubic-foot to NMFC freight class."""
    for min_pcf, freight_class in _DENSITY_BANDS:
        if pcf >= min_pcf:
            return freight_class
    return 400


def apply_height_rule(height_in: float, skip_seventy_five_inch_rule: bool) -> tuple[float, bool]:
    """XPO 75–95 inch rule: bump height to 96 unless skipped."""
    if skip_seventy_five_inch_rule:
        return height_in, False
    if 75 <= height_in <= 95:
        return 96.0, True
    return height_in, False


def line_cubic_feet(length_in: float, width_in: float, height_in: float, pallets: int) -> float:
    return (length_in * width_in * height_in / CUBIC_INCHES_PER_CU_FT) * pallets


def calculate_shipment(
    shipment_id: str,
    rows: list[dict[str, Any]],
    *,
    skip_seventy_five_inch_rule: bool = False,
) -> ShipmentResult:
    """Calculate aggregate density and freight class for one shipment."""
    if not rows:
        raise FreightClassCalculatorError(f"Shipment '{shipment_id}' has no pallet rows.")

    line_items: list[ShipmentLineItem] = []
    total_weight = 0.0
    total_cube = 0.0
    any_height_rule = False

    for index, row in enumerate(rows, start=1):
        try:
            pallets = int(row["pallets"])
            weight = float(row["weight"])
            length = float(row["length"])
            width = float(row["width"])
            height = float(row["height"])
        except (KeyError, TypeError, ValueError) as exc:
            raise FreightClassCalculatorError(
                f"Shipment '{shipment_id}' row {index}: invalid numeric values."
            ) from exc

        if pallets < 1:
            raise FreightClassCalculatorError(
                f"Shipment '{shipment_id}' row {index}: pallet count must be at least 1."
            )
        if weight <= 0 or length <= 0 or width <= 0 or height <= 0:
            raise FreightClassCalculatorError(
                f"Shipment '{shipment_id}' row {index}: weight and dimensions must be positive."
            )

        adjusted_height, rule_applied = apply_height_rule(height, skip_seventy_five_inch_rule)
        cube = line_cubic_feet(length, width, adjusted_height, pallets)
        any_height_rule = any_height_rule or rule_applied

        item = ShipmentLineItem(
            pallets=pallets,
            weight_lbs=weight,
            length_in=length,
            width_in=width,
            height_in=height,
            adjusted_height_in=adjusted_height if rule_applied else None,
            height_rule_applied=rule_applied,
            cubic_feet=cube,
        )
        line_items.append(item)
        total_weight += weight
        total_cube += cube

    if total_cube <= 0:
        raise FreightClassCalculatorError(f"Shipment '{shipment_id}' has zero volume.")

    density = total_weight / total_cube
    return ShipmentResult(
        shipment_id=shipment_id,
        line_items=line_items,
        total_weight_lbs=total_weight,
        total_cubic_feet=total_cube,
        density_pcf=density,
        freight_class=density_to_freight_class(density),
        height_rule_applied=any_height_rule,
    )


def calculate_from_grouped_shipments(
    grouped: dict[str, list[dict[str, Any]]],
    *,
    skip_seventy_five_inch_rule: bool = False,
) -> CalculationResult:
    if not grouped:
        raise FreightClassCalculatorError("No shipment data to calculate.")
    shipments = [
        calculate_shipment(sid, rows, skip_seventy_five_inch_rule=skip_seventy_five_inch_rule)
        for sid, rows in grouped.items()
    ]
    return CalculationResult(shipments=shipments)


def calculate_manual(
    line_items: list[dict[str, Any]],
    *,
    shipment_id: str = "Manual Entry",
    skip_seventy_five_inch_rule: bool = False,
) -> CalculationResult:
    return calculate_from_grouped_shipments(
        {shipment_id: line_items},
        skip_seventy_five_inch_rule=skip_seventy_five_inch_rule,
    )


def _normalize_header(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    return _HEADER_ALIASES.get(text)


def _cell_is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _row_is_empty(values: tuple[Any, ...]) -> bool:
    return all(_cell_is_blank(v) for v in values)


def parse_excel_shipments(file_bytes: bytes) -> dict[str, list[dict[str, Any]]]:
    """Parse pallet-dimension workbook grouped by shipment ID (forward-filled)."""
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise FreightClassCalculatorError("Could not read Excel file.") from exc

    sheet = workbook.active
    if sheet is None:
        raise FreightClassCalculatorError("Excel file has no worksheets.")

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise FreightClassCalculatorError("Excel file is missing a header row.")

    column_map: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _normalize_header(cell)
        if key and key not in column_map:
            column_map[key] = idx

    required = {"pallets", "weight", "length", "width", "height"}
    missing = required - set(column_map)
    if missing:
        labels = ", ".join(sorted(missing))
        raise FreightClassCalculatorError(
            f"Missing required column(s): {labels}. "
            "Expected: Shipment ID, Pallet Number, Pallet Count, Weight, Length, Width, Height "
            "(legacy 'Pallets' column is also accepted as Pallet Count)."
        )

    grouped: dict[str, list[dict[str, Any]]] = {}
    current_shipment: str | None = None
    row_num = 1

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_num += 1
        if _row_is_empty(row):
            continue

        shipment_cell = row[column_map["shipment_id"]] if "shipment_id" in column_map else None
        if not _cell_is_blank(shipment_cell):
            current_shipment = str(shipment_cell).strip()
        if not current_shipment:
            raise FreightClassCalculatorError(
                f"Row {row_num}: Shipment ID is required on the first row of each shipment group."
            )

        try:
            pallets_raw = row[column_map["pallets"]]
            weight_raw = row[column_map["weight"]]
            length_raw = row[column_map["length"]]
            width_raw = row[column_map["width"]]
            height_raw = row[column_map["height"]]
            if any(_cell_is_blank(v) for v in (pallets_raw, weight_raw, length_raw, width_raw, height_raw)):
                raise FreightClassCalculatorError(
                    f"Row {row_num} ({current_shipment}): all pallet fields must be filled."
                )
            parsed = {
                "pallets": int(float(pallets_raw)),
                "weight": float(weight_raw),
                "length": float(length_raw),
                "width": float(width_raw),
                "height": float(height_raw),
            }
        except FreightClassCalculatorError:
            raise
        except (TypeError, ValueError) as exc:
            raise FreightClassCalculatorError(
                f"Row {row_num} ({current_shipment}): invalid numeric value."
            ) from exc

        grouped.setdefault(current_shipment, []).append(parsed)

    if not grouped:
        raise FreightClassCalculatorError("No pallet rows found in the uploaded file.")

    return grouped


def calculate_from_excel(
    file_bytes: bytes,
    *,
    skip_seventy_five_inch_rule: bool = False,
) -> CalculationResult:
    grouped = parse_excel_shipments(file_bytes)
    return calculate_from_grouped_shipments(
        grouped,
        skip_seventy_five_inch_rule=skip_seventy_five_inch_rule,
    )


def build_results_workbook(result: CalculationResult) -> bytes:
    """Build summary + detail Excel export from calculation results."""
    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    accent_fill = PatternFill(start_color="81B81D", end_color="81B81D", fill_type="solid")

    summary_headers = [
        "Shipment ID",
        "Total Pallets",
        "Total Weight (lbs)",
        "Total Volume (ft³)",
        "Density (lb/ft³)",
        "Freight Class",
        "75\" Rule Applied",
    ]
    ws_summary.append(summary_headers)
    for col in range(1, len(summary_headers) + 1):
        cell = ws_summary.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for shipment in result.shipments:
        total_pallets = sum(item.pallets for item in shipment.line_items)
        ws_summary.append(
            [
                shipment.shipment_id,
                total_pallets,
                round(shipment.total_weight_lbs, 2),
                round(shipment.total_cubic_feet, 4),
                round(shipment.density_pcf, 4),
                shipment.freight_class,
                "Yes" if shipment.height_rule_applied else "No",
            ]
        )

    for row in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row):
        class_cell = row[5]
        class_cell.font = Font(bold=True)
        class_cell.fill = accent_fill
        class_cell.alignment = Alignment(horizontal="center")

    for col in ws_summary.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws_summary.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    # Detail sheet
    ws_detail = wb.create_sheet("Line Items")
    detail_headers = [
        "Shipment ID",
        "Pallet Number",
        "Pallet Count",
        "Weight (lbs)",
        "Length (in)",
        "Width (in)",
        "Height (in)",
        "Adj. Height (in)",
        "Line Volume (ft³)",
        "75\" Rule",
    ]
    ws_detail.append(detail_headers)
    for col in range(1, len(detail_headers) + 1):
        cell = ws_detail.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for shipment in result.shipments:
        for index, item in enumerate(shipment.line_items, start=1):
            ws_detail.append(
                [
                    shipment.shipment_id,
                    f"#{index}",
                    item.pallets,
                    round(item.weight_lbs, 2),
                    round(item.length_in, 2),
                    round(item.width_in, 2),
                    round(item.height_in, 2),
                    round(item.adjusted_height_in, 2) if item.adjusted_height_in else "",
                    round(item.cubic_feet, 4),
                    "Yes" if item.height_rule_applied else "",
                ]
            )

    for col in ws_detail.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws_detail.column_dimensions[col[0].column_letter].width = min(max_len + 3, 24)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pallet Dims"
    # Pallet Number = line label (#1, #2). Pallet Count = identical-pallet multiplier.
    headers = [
        "Shipment ID",
        "Pallet Number",
        "Pallet Count",
        "Weight",
        "Length",
        "Width",
        "Height",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    ws.append(["FBA19EXAMPLE1", "#1", 1, 500, 48, 40, 36])
    ws.append([None, "#2", 2, 350, 48, 40, 52])
    ws.append(["FBA19EXAMPLE2", "#1", 1, 458, 48, 40, 56])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
