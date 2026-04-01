"""CSV generation service for price alerts."""
import pandas as pd
import io
import logging
import re
import httpx
from typing import List, Dict, Any, Optional
from datetime import datetime
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class CSVGenerator:
    """Generates CSV files from price alert data."""
    
    @staticmethod
    def fetch_buy_box_price_from_amazon(amazon_url: str) -> Optional[float]:
        """
        Fetch the buy box price directly from Amazon URL.
        
        Args:
            amazon_url: Amazon product URL (e.g., https://www.amazon.com/dp/B0DQRGZ32G)
            
        Returns:
            Buy box price as float, or None if unable to fetch
        """
        if not amazon_url or amazon_url == "N/A":
            return None
        
        try:
            import json
            
            # Use proper headers to avoid being blocked
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.5",
                "Accept-Encoding": "gzip, deflate, br",
                "Connection": "keep-alive",
                "Upgrade-Insecure-Requests": "1",
                "Referer": "https://www.amazon.com/",
            }
            
            # Use httpx sync client with timeout and better error handling
            with httpx.Client(timeout=15.0, follow_redirects=True) as client:
                logger.info(f"Fetching buy box price from Amazon URL: {amazon_url}")
                response = client.get(amazon_url, headers=headers)
                response.raise_for_status()
                html_content = response.text
                
                if len(html_content) < 1000:
                    logger.warning(f"Amazon returned very short response, might be blocked: {len(html_content)} chars")
                
                # Try multiple methods to extract the price
                price = None
                
                # Method 1: Parse JSON data from script tags (most reliable)
                # Amazon embeds product data in window.ue_backflow_data or similar
                try:
                    # Look for JSON data in script tags
                    script_json_patterns = [
                        r'var\s+ue_backflow_data\s*=\s*({.*?});',
                        r'window\.ue_backflow_data\s*=\s*({.*?});',
                        r'var\s+obj\s*=\s*jQuery\.parseJSON\(\'({.*?})\'\)',
                        r'"priceToPay".*?"amount":(\d+\.?\d*)',
                        r'"buyBoxPrice".*?"amount":(\d+\.?\d*)',
                        r'"displayPrice".*?"amount":(\d+\.?\d*)',
                    ]
                    
                    for pattern in script_json_patterns:
                        matches = re.findall(pattern, html_content, re.DOTALL)
                        for match in matches:
                            try:
                                # Try to parse as JSON if it looks like JSON
                                if match.strip().startswith('{'):
                                    data = json.loads(match)
                                    # Look for price in various nested structures
                                    price = CSVGenerator._extract_price_from_json(data)
                                    if price:
                                        logger.info(f"Found price from JSON script tag: ${price:.2f}")
                                        break
                                else:
                                    # Direct number match
                                    price_str = match.replace(',', '').strip()
                                    price = float(price_str)
                                    if price > 0:
                                        logger.info(f"Found price from script pattern: ${price:.2f}")
                                        break
                            except (json.JSONDecodeError, ValueError, KeyError):
                                continue
                        if price:
                            break
                except Exception as e:
                    logger.debug(f"Error parsing JSON from script tags: {e}")
                
                # Method 2: Look for price in the buy box section specifically (Add to Cart box)
                # This targets the price shown in the buy box area where "Add to cart" button is
                if price is None:
                    # Extract the buy box section first - look for the "Add to Cart" area
                    buy_box_patterns = [
                        r'<div[^>]*id="buybox"[^>]*>.*?</div>',
                        r'<div[^>]*id="desktop_buybox"[^>]*>.*?</div>',
                        r'<div[^>]*class="[^"]*buybox[^"]*"[^>]*>.*?</div>',
                        r'<div[^>]*id="addToCart"[^>]*>.*?</div>',
                        r'<div[^>]*class="[^"]*a-section[^"]*"[^>]*>.*?Add to cart.*?</div>',
                    ]
                    
                    buy_box_html = ""
                    for pattern in buy_box_patterns:
                        matches = re.findall(pattern, html_content, re.DOTALL | re.IGNORECASE)
                        if matches:
                            buy_box_html = matches[0]
                            logger.debug(f"Found buy box section with {len(buy_box_html)} characters")
                            break
                    
                    if buy_box_html:
                        # Look for price in buy box section - prioritize a-price-whole and a-price-symbol
                        # Amazon typically shows price as: <span class="a-price-whole">219</span><span class="a-price-symbol">^</span><span class="a-price-fraction">95</span>
                        price_patterns = [
                            # Pattern for split price: whole + fraction (e.g., 219^95 = $219.95)
                            r'<span[^>]*class="[^"]*a-price-whole[^"]*"[^>]*>(\d+)</span>.*?<span[^>]*class="[^"]*a-price-fraction[^"]*"[^>]*>(\d+)</span>',
                            # Pattern for full price in a-offscreen (screen reader text)
                            r'<span[^>]*class="[^"]*a-offscreen[^"]*"[^>]*>\$(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)</span>',
                            # Pattern for price-whole only
                            r'<span[^>]*class="[^"]*a-price-whole[^"]*"[^>]*>(\d+)</span>',
                            # Pattern for any price with dollar sign
                            r'\$(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
                        ]
                        
                        for pattern in price_patterns:
                            matches = re.findall(pattern, buy_box_html, re.DOTALL | re.IGNORECASE)
                            if matches:
                                try:
                                    # Handle split price format (whole + fraction)
                                    if isinstance(matches[0], tuple) and len(matches[0]) == 2:
                                        whole_part = matches[0][0].replace(',', '').strip()
                                        fraction_part = matches[0][1].strip()
                                        price_str = f"{whole_part}.{fraction_part}"
                                        price = float(price_str)
                                    else:
                                        price_str = str(matches[0]).replace(',', '').replace('$', '').strip()
                                        price = float(price_str)
                                    
                                    if price > 0 and price < 100000:  # Reasonable price range
                                        logger.info(f"Found price from buy box section: ${price:.2f}")
                                        break
                                except (ValueError, IndexError, TypeError):
                                    continue
                
                # Method 3: Look for price in HTML elements (common selectors)
                if price is None:
                    price_patterns = [
                        r'<span[^>]*id="priceblock_[^"]*"[^>]*>\$([^<]+)</span>',
                        r'<span[^>]*class="[^"]*a-price-whole[^"]*"[^>]*>([^<]+)</span>',
                        r'<span[^>]*class="[^"]*a-price[^"]*"[^>]*>.*?<span[^>]*class="[^"]*a-offscreen[^"]*"[^>]*>\$([^<]+)</span>',
                        r'<span[^>]*class="[^"]*a-price[^"]*"[^>]*>.*?\$(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
                        r'id="priceblock_[^"]*"[^>]*>\s*\$(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
                    ]
                    
                    for pattern in price_patterns:
                        matches = re.findall(pattern, html_content, re.DOTALL | re.IGNORECASE)
                        if matches:
                            try:
                                price_str = matches[0].replace(',', '').replace('$', '').strip()
                                price = float(price_str)
                                if price > 0:
                                    logger.info(f"Found price from HTML pattern: ${price:.2f}")
                                    break
                            except (ValueError, IndexError):
                                continue
                
                # Method 4: Look for price in meta tags or data attributes
                if price is None:
                    meta_patterns = [
                        r'<meta[^>]*property="product:price:amount"[^>]*content="([^"]+)"',
                        r'data-a-price="([^"]+)"',
                        r'data-price="([^"]+)"',
                    ]
                    
                    for pattern in meta_patterns:
                        matches = re.findall(pattern, html_content, re.IGNORECASE)
                        if matches:
                            try:
                                price_str = matches[0].replace(',', '').replace('$', '').strip()
                                price = float(price_str)
                                if price > 0:
                                    logger.info(f"Found price from meta/data pattern: ${price:.2f}")
                                    break
                            except (ValueError, IndexError):
                                continue
                
                if price is not None and price > 0:
                    logger.info(f"Successfully extracted buy box price ${price:.2f} from {amazon_url}")
                    return price
                else:
                    logger.warning(f"Could not extract valid price from {amazon_url}. HTML length: {len(html_content)}")
                    return None
                    
        except httpx.HTTPError as e:
            logger.warning(f"HTTP error fetching Amazon URL {amazon_url}: {e}")
            return None
        except Exception as e:
            logger.warning(f"Error fetching buy box price from {amazon_url}: {e}", exc_info=True)
            return None
    
    @staticmethod
    def _extract_price_from_json(data: dict, depth: int = 0) -> Optional[float]:
        """Recursively extract price from nested JSON structure."""
        if depth > 5:  # Prevent infinite recursion
            return None
        
        # Common price field names
        price_fields = ['price', 'buyBoxPrice', 'displayPrice', 'priceToPay', 'amount', 'value']
        
        for key, value in data.items():
            if key.lower() in [f.lower() for f in price_fields]:
                try:
                    if isinstance(value, (int, float)):
                        price = float(value)
                        # If price seems to be in cents (very large number), convert
                        if price > 10000:
                            price = price / 100.0
                        if price > 0:
                            return price
                    elif isinstance(value, str):
                        # Try to extract number from string
                        price_str = re.sub(r'[^\d.]', '', value)
                        if price_str:
                            price = float(price_str)
                            if price > 0:
                                return price
                    elif isinstance(value, dict):
                        # Look for nested amount or value
                        if 'amount' in value:
                            amount = value['amount']
                            if isinstance(amount, (int, float)):
                                price = float(amount)
                                if price > 10000:
                                    price = price / 100.0
                                if price > 0:
                                    return price
                except (ValueError, TypeError):
                    pass
            
            # Recursively search nested dictionaries
            if isinstance(value, dict):
                result = CSVGenerator._extract_price_from_json(value, depth + 1)
                if result:
                    return result
            elif isinstance(value, list):
                for item in value:
                    if isinstance(item, dict):
                        result = CSVGenerator._extract_price_from_json(item, depth + 1)
                        if result:
                            return result
        
        return None
    
    @staticmethod
    def generate_price_alerts_csv(price_alerts: List[Dict[str, Any]]) -> bytes:
        """
        Generate CSV file from price alerts data.
        
        Args:
            price_alerts: List of price alert dictionaries
            
        Returns:
            CSV file as bytes
        """
        if not price_alerts:
            # Return empty CSV with headers
            df = pd.DataFrame(columns=[
                "UPC",
                "Seller Name",
                "Current Price",
                "Historical Price",
                "Price Change %",
                "Detected At"
            ])
        else:
            # Prepare data for DataFrame
            csv_data = []
            for alert in price_alerts:
                csv_data.append({
                    "UPC": alert.get("upc", ""),
                    "Seller Name": alert.get("seller_name", ""),
                    "Current Price": float(alert.get("current_price", 0)) if alert.get("current_price") else "",
                    "Historical Price": float(alert.get("historical_price", 0)) if alert.get("historical_price") else "",
                    "Price Change %": float(alert.get("price_change_percent", 0)) if alert.get("price_change_percent") else "",
                    "Detected At": alert.get("detected_at", ""),
                })
            
            df = pd.DataFrame(csv_data)
        
        # Generate CSV bytes
        csv_buffer = io.BytesIO()
        df.to_csv(csv_buffer, index=False, encoding="utf-8")
        csv_bytes = csv_buffer.getvalue()
        csv_buffer.close()
        
        return csv_bytes
    
    @staticmethod
    def generate_csv_filename(job_name: str = "keepa_report", extension: str = "xlsx") -> str:
        """
        Generate filename with timestamp.
        
        Args:
            job_name: Base name for the file
            extension: File extension (default: xlsx for Excel)
            
        Returns:
            Filename string
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # Sanitize job_name for filename
        safe_name = "".join(c for c in job_name if c.isalnum() or c in (" ", "-", "_")).strip()
        safe_name = safe_name.replace(" ", "_")
        filename = f"{safe_name}_{timestamp}.{extension}"
        return filename
    
    @staticmethod
    def convert_alerts_to_csv_format(alerts: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Convert price alerts from database format to CSV format.
        
        Args:
            alerts: List of price alert dictionaries from database
            
        Returns:
            List of dictionaries formatted for CSV generation
        """
        return [
            {
                "upc": alert["upc"],
                "seller_name": alert.get("seller_name"),
                "current_price": alert.get("current_price"),
                "historical_price": alert.get("historical_price"),
                "price_change_percent": alert.get("price_change_percent"),
                "detected_at": alert.get("detected_at"),
            }
            for alert in alerts
        ]
    
    @staticmethod
    def extract_keepa_product_data(keepa_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Extract product information from Keepa data.
        
        Args:
            keepa_data: Raw Keepa API response
            
        Returns:
            Dict with extracted product data
        """
        if not keepa_data or not isinstance(keepa_data, dict):
            return {}
        
        products = keepa_data.get("products", [])
        if not products or len(products) == 0:
            return {}
        
        product = products[0]
        stats = product.get("stats", {})
        current_sellers = product.get("current_sellers", [])
        
        # Get buy box seller ID from stats
        buy_box_seller_id = stats.get("buyBoxSellerId", None)
        buy_box_price = None
        buy_box_seller_name = None
        
        # FIRST PRIORITY: Get buy box price from stats.buyBoxPrice
        # This is the actual buy box price shown on Amazon (in cents)
        buy_box_price = (
            stats.get("buyBoxPrice") or 
            stats.get("buyBoxPriceNew") or 
            stats.get("current") or
            None
        )
        
        # Convert from Keepa price format (cents) to dollars
        # Keepa stores all prices in cents, so always divide by 100
        if buy_box_price is not None:
            try:
                # Handle negative prices (returns, etc.) - only use positive prices
                buy_box_price_cents = float(buy_box_price)
                if buy_box_price_cents > 0:
                    buy_box_price = buy_box_price_cents / 100.0
                else:
                    # Negative or zero price is invalid, set to None
                    buy_box_price = None
                    logger.debug(f"Invalid buy box price from Keepa stats: {buy_box_price_cents} cents")
            except (TypeError, ValueError):
                buy_box_price = None
        
        # Find buy box seller name from current_sellers
        if buy_box_seller_id and current_sellers:
            for seller in current_sellers:
                if seller.get("sellerId") == buy_box_seller_id:
                    buy_box_seller_name = seller.get("sellerName", "")
                    # If we don't have buy box price from stats, use seller's price
                    if buy_box_price is None:
                        seller_price = seller.get("price")
                        if seller_price is not None:
                            try:
                                seller_price_cents = float(seller_price)
                                if seller_price_cents > 0:
                                    buy_box_price = seller_price_cents / 100.0
                            except (TypeError, ValueError):
                                pass
                    break
        
        # Fallback: If buy box seller not found by ID, try to find Amazon or the first seller
        if not buy_box_seller_name and current_sellers:
            for seller in current_sellers:
                seller_name = seller.get("sellerName", "")
                if "amazon" in seller_name.lower() or seller.get("isFBA", False):
                    buy_box_seller_name = seller_name
                    buy_box_seller_id = buy_box_seller_id or seller.get("sellerId")
                    # If we still don't have buy box price, use this seller's price
                    if buy_box_price is None:
                        seller_price = seller.get("price")
                        if seller_price is not None:
                            try:
                                seller_price_cents = float(seller_price)
                                if seller_price_cents > 0:
                                    buy_box_price = seller_price_cents / 100.0
                            except (TypeError, ValueError):
                                pass
                    break
            
            # If still no buy box seller, use first seller
            if not buy_box_seller_name and current_sellers:
                first_seller = current_sellers[0]
                buy_box_seller_name = first_seller.get("sellerName", "")
                buy_box_seller_id = buy_box_seller_id or first_seller.get("sellerId")
                # If we still don't have buy box price, use first seller's price
                if buy_box_price is None:
                    seller_price = first_seller.get("price")
                    if seller_price is not None:
                        try:
                            seller_price_cents = float(seller_price)
                            if seller_price_cents > 0:
                                buy_box_price = seller_price_cents / 100.0
                        except (TypeError, ValueError):
                            pass
        
        # Get current Amazon price (buy box price or lowest price)
        current_amazon_price = buy_box_price
        if current_amazon_price is None and current_sellers:
            # Get lowest price from sellers
            prices = []
            for seller in current_sellers:
                price = seller.get("price")
                if price is not None:
                    try:
                        # Convert from cents to dollars (Keepa stores prices in cents)
                        price_dollars = float(price) / 100.0
                        if price_dollars > 0:
                            prices.append(price_dollars)
                    except (TypeError, ValueError):
                        pass
            if prices:
                current_amazon_price = min(prices)
        
        # Format seller ID for output (Keepa returns int or str)
        buy_box_seller_id_display = str(buy_box_seller_id) if buy_box_seller_id is not None else ""
        return {
            "asin": product.get("asin", ""),
            "title": product.get("title", ""),
            "brand": product.get("brand", ""),
            "buy_box_price": buy_box_price,
            "buy_box_seller_name": buy_box_seller_name or "",
            "buy_box_seller_id": buy_box_seller_id_display,
            "current_amazon_price": current_amazon_price,
        }

    @staticmethod
    def _seller_offers_from_keepa(keepa_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        All seller offers from current_sellers. Keepa prices are in cents.

        One entry per row in Keepa's list (duplicate sellerIds can appear — each becomes a row).
        Each dict: price (float dollars), seller_name (str), seller_id (str).
        """
        offers: List[Dict[str, Any]] = []
        if not keepa_data or not isinstance(keepa_data, dict):
            return offers
        products = keepa_data.get("products", [])
        if not products:
            return offers
        for seller in products[0].get("current_sellers", []):
            raw = seller.get("price")
            if raw is None:
                continue
            try:
                cents = float(raw)
                if cents <= 0:
                    continue
                dollars = cents / 100.0
                sn = (seller.get("sellerName") or "").strip()
                sid = seller.get("sellerId")
                offers.append({
                    "price": dollars,
                    "seller_name": sn,
                    "seller_id": str(sid) if sid is not None else "",
                })
            except (TypeError, ValueError):
                continue
        return offers

    @staticmethod
    def generate_comprehensive_report_csv(
        processed_items: List[Dict[str, Any]],
        price_alerts_by_upc: Dict[str, List[Dict[str, Any]]],
        map_prices_by_upc: Dict[str, Decimal],
        seller_name_map: Optional[Dict[str, str]] = None
    ) -> tuple:
        """
        Generate comprehensive CSV report matching the spreadsheet format.
        
        Args:
            processed_items: List of upc_batch_items with keepa_data
            price_alerts_by_upc: Dict mapping UPC to list of price alerts
            map_prices_by_upc: Dict mapping UPC to MAP price
            seller_name_map: Optional dict mapping seller_id to seller_name for display
            
        Returns:
            Tuple of (Excel file as bytes, number of off-price rows — one per seller below MAP)
        """
        csv_data = []
        off_price_count = 0
        seller_name_map = seller_name_map or {}

        for item in processed_items:
            upc = item.get("upc", "")
            keepa_data = item.get("keepa_data", {})

            product_data = CSVGenerator.extract_keepa_product_data(keepa_data)

            map_price = map_prices_by_upc.get(upc)
            try:
                msrp = float(map_price) if map_price else None
            except (TypeError, ValueError):
                msrp = None

            asin = product_data.get("asin", "")
            amazon_url = f"https://www.amazon.com/dp/{asin}" if asin else "N/A"

            # Format UPC (handle scientific notation)
            upc_display = upc
            try:
                if "E+" in str(upc) or "e+" in str(upc):
                    upc_display = f"{float(upc):.0f}"
            except Exception:
                pass

            base_title = product_data.get("title", "")
            base_brand = product_data.get("brand", "")

            def append_row(
                reference_price: float,
                seller_display: str,
            ) -> None:
                nonlocal off_price_count
                off_price_count += 1
                try:
                    pdiff = float(msrp) - float(reference_price)
                    price_difference_display = f"${pdiff:.2f}"
                except (TypeError, ValueError):
                    price_difference_display = "$0.00"
                if msrp is not None and float(msrp) > 0:
                    try:
                        discount_percent = (
                            (float(msrp) - float(reference_price)) / float(msrp)
                        ) * 100
                        discount_display = f"{discount_percent:.2f}%"
                    except (TypeError, ValueError, ZeroDivisionError):
                        discount_display = "N/A"
                else:
                    discount_display = "N/A"
                try:
                    cur_disp = (
                        f"${float(reference_price):.2f}"
                        if reference_price is not None
                        else "N/A"
                    )
                except (TypeError, ValueError):
                    cur_disp = "N/A"
                try:
                    seller_price_disp = f"${float(reference_price):.2f}"
                except (TypeError, ValueError):
                    seller_price_disp = "N/A"
                csv_data.append({
                    "UPC": upc_display,
                    "ASIN": asin,
                    "Product Title": base_title,
                    "Brand": base_brand,
                    "Off Price Listing": "Off Price",
                    "MSRP": f"${msrp:.2f}" if msrp else "N/A",
                    "Current Amazon Price": cur_disp,
                    "Price Difference": price_difference_display,
                    "Seller Offer Price": seller_price_disp,
                    "Seller": seller_display,
                    "Discount %": discount_display,
                    "Amazon URL": amazon_url,
                    "_is_off_price": True,
                })

            if msrp is None:
                continue

            offers = CSVGenerator._seller_offers_from_keepa(keepa_data)

            for off in offers:
                ref = off["price"]
                if not (float(msrp) > float(ref)):
                    continue
                sid = off.get("seller_id") or ""
                sname = off.get("seller_name") or ""
                if sname:
                    seller_disp = sname
                elif sid and seller_name_map:
                    seller_disp = seller_name_map.get(sid, sid)
                elif sid:
                    seller_disp = sid
                else:
                    seller_disp = "N/A"
                append_row(ref, seller_disp)

            # No seller rows in Keepa: fallback to buy box vs MAP (single row)
            if not offers:
                buy_box_only = product_data.get("buy_box_price")
                if buy_box_only is not None and float(msrp) > float(buy_box_only):
                    bb_name = product_data.get("buy_box_seller_name", "")
                    bb_id = product_data.get("buy_box_seller_id", "")
                    if bb_name:
                        seller_disp = bb_name
                    elif bb_id and seller_name_map:
                        seller_disp = seller_name_map.get(bb_id, bb_id)
                    elif bb_id:
                        seller_disp = bb_id
                    else:
                        seller_disp = "N/A"
                    append_row(float(buy_box_only), seller_disp)
                elif buy_box_only is None:
                    logger.warning(
                        f"No seller or buy box price found in Keepa data for UPC {upc}"
                    )

        # Create DataFrame with all columns
        df = pd.DataFrame(csv_data)
        
        # Ensure all columns are present even if empty
        required_columns = [
            "UPC", "ASIN", "Product Title", "Brand", "Off Price Listing",
            "MSRP", "Current Amazon Price", "Price Difference",
            "Seller Offer Price", "Seller", "Discount %", "Amazon URL"
        ]
        
        for col in required_columns + ["_is_off_price"]:
            if col not in df.columns:
                df[col] = "" if col != "_is_off_price" else False
        
        # Reorder columns (exclude internal _is_off_price column)
        df = df[required_columns + ["_is_off_price"]]
        
        # Generate Excel file with formatting instead of CSV
        excel_buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Price Report"
        
        # Write headers
        headers = required_columns
        ws.append(headers)
        
        # Style header row
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
        
        # Write data rows
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        white_font = Font(color="FFFFFFFF", bold=True)
        
        for idx, row in df.iterrows():
            row_data = [row[col] for col in required_columns]
            ws.append(row_data)
            
            # Highlight "Off Price Listing" column (Column E, index 4) in red if it's "Off Price"
            if row.get("_is_off_price", False):
                # Column E is index 5 (1-based) in Excel, row is idx + 2 (header + 1-based)
                off_price_cell = ws.cell(row=idx + 2, column=5)  # Column E = 5
                off_price_cell.fill = red_fill
                off_price_cell.font = white_font
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        wb.save(excel_buffer)
        excel_bytes = excel_buffer.getvalue()
        excel_buffer.close()
        
        return excel_bytes, off_price_count

