"""Parse UPC / DIMS sheets from the catalog spreadsheet."""
from __future__ import annotations

import io
import logging
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from app.services.catalog_upc_dims_headers import (
    DIMS_HEADERS,
    DIMS_REQUIRED_HEADER,
    DIMS_SHEET_NAME,
    UPC_HEADERS,
    UPC_REQUIRED_HEADER,
    UPC_SHEET_NAME,
)

logger = logging.getLogger(__name__)


def _cell_str(raw: Any) -> str:
    """Normalize Excel cell values (avoid float UPCs like 190038644083.0)."""
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return "TRUE" if raw else "FALSE"
    if isinstance(raw, datetime):
        if raw.hour == 0 and raw.minute == 0 and raw.second == 0 and raw.microsecond == 0:
            return raw.date().isoformat()
        return raw.isoformat(sep=" ", timespec="seconds")
    if isinstance(raw, date):
        return raw.isoformat()
    if isinstance(raw, int):
        return str(raw)
    if isinstance(raw, float):
        if raw.is_integer():
            return str(int(raw))
        return format(raw, ".15g").strip()
    return str(raw).strip()


def _normalize_header(cell: Any) -> str:
    return _cell_str(cell).strip()


def _find_sheet(wb, preferred: str):
    for name in wb.sheetnames:
        if name.strip().upper() == preferred.upper():
            return wb[name]
    return wb[wb.sheetnames[0]] if wb.sheetnames else None


def _header_index_map(
    header_row: Sequence[Any],
    expected: Sequence[str],
    required: str,
) -> Dict[str, int]:
    found: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        name = _normalize_header(cell)
        if not name:
            continue
        if name in found:
            continue
        found[name] = idx

    if required not in found:
        raise ValueError(
            f'Spreadsheet must include a "{required}" column. '
            f"Download the template and match the exact headers."
        )

    # Prefer exact expected headers; accept extras but require all expected present.
    missing = [h for h in expected if h not in found]
    if missing:
        preview = ", ".join(missing[:8])
        more = f" (+{len(missing) - 8} more)" if len(missing) > 8 else ""
        raise ValueError(
            f"Missing required column(s): {preview}{more}. "
            "Use the downloadable template (exact headers from UPC DIMS.xlsx)."
        )
    return {h: found[h] for h in expected}


def _parse_sheet_rows(
    filename: str,
    content: bytes,
    sheet_name: str,
    expected_headers: Sequence[str],
    required_header: str,
    key_header: str,
) -> Tuple[List[Dict[str, str]], int]:
    lower = (filename or "").lower()
    if not lower.endswith((".xlsx", ".xlsm", ".xls")):
        raise ValueError("Upload an .xlsx file matching the UPC / DIMS template.")

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = _find_sheet(wb, sheet_name)
        if sheet is None:
            raise ValueError("Workbook has no sheets.")
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], 0

        mapping = _header_index_map(header_row, expected_headers, required_header)
        valid: List[Dict[str, str]] = []
        invalid = 0
        for row in rows_iter:
            row_data: Dict[str, str] = {}
            any_value = False
            for header in expected_headers:
                idx = mapping[header]
                value = _cell_str(row[idx]) if idx < len(row) else ""
                if value:
                    any_value = True
                row_data[header] = value
            if not any_value:
                continue
            key = (row_data.get(key_header) or "").strip()
            if not key:
                invalid += 1
                continue
            valid.append(row_data)
        return valid, invalid
    finally:
        wb.close()


def parse_upc_spreadsheet(filename: str, content: bytes) -> Tuple[List[Dict[str, str]], int]:
    return _parse_sheet_rows(
        filename,
        content,
        UPC_SHEET_NAME,
        UPC_HEADERS,
        UPC_REQUIRED_HEADER,
        UPC_REQUIRED_HEADER,
    )


def parse_dims_spreadsheet(filename: str, content: bytes) -> Tuple[List[Dict[str, str]], int]:
    return _parse_sheet_rows(
        filename,
        content,
        DIMS_SHEET_NAME,
        DIMS_HEADERS,
        DIMS_REQUIRED_HEADER,
        DIMS_REQUIRED_HEADER,
    )


def dedupe_by_key(rows: List[Dict[str, str]], key_header: str) -> List[Dict[str, str]]:
    """Last row wins for duplicate keys within one import file."""
    seen: Dict[str, Dict[str, str]] = {}
    for row in rows:
        key = (row.get(key_header) or "").strip()
        if key:
            seen[key] = row
    return list(seen.values())


def upc_row_to_record(row_data: Dict[str, str]) -> Dict[str, Any]:
    return {
        "upc_code": (row_data.get("UPC Code") or "").strip(),
        "scs": row_data.get("S/C/S") or "",
        "vendor_name": row_data.get("Vendor Name") or "",
        "display_name": row_data.get("Display Name") or "",
        "netsuite_style_name": row_data.get("Netsuite Style Name") or "",
        "status": row_data.get("STATUS") or "",
        "row_data": row_data,
    }


def dims_row_to_record(row_data: Dict[str, str]) -> Dict[str, Any]:
    return {
        "upc_number": (row_data.get("UPC #") or "").strip(),
        "sku": row_data.get("SKU") or "",
        "brand": row_data.get("Brand") or "",
        "description": row_data.get("Description") or "",
        "current_season": row_data.get("Current Season") or "",
        "item_status": row_data.get("Item Status") or "",
        "row_data": row_data,
    }
