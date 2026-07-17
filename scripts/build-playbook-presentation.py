#!/usr/bin/env python3
"""Generate HTML presentation from Master Sheet Data Analyst's Playbook.xlsx"""
from html import escape
from pathlib import Path
from openpyxl import load_workbook

XLSX = Path(r"c:\Users\Administrator\Desktop\Master Sheet Data Analyst's Playbook.xlsx")
OUT = Path(__file__).resolve().parent.parent / "docs" / "master-sheet-data-analyst-playbook.html"


def parse_playbook():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb.active
    sections = []
    current = None

    for r in range(2, ws.max_row + 1):
        p = ws.cell(r, 1).value
        c2 = ws.cell(r, 2).value
        c3 = ws.cell(r, 3).value
        p = str(p).strip() if p else ""
        c2 = str(c2).strip() if c2 else ""
        c3 = str(c3).strip() if c3 else ""
        if not p and not c2 and not c3:
            continue

        if p and (not c2 or (len(p) < 90 and not p.endswith("?") and "?" not in p[:20])):
            # New section header (topic in col A)
            if p.startswith("amzn."):
                if current:
                    if c2:
                        current["facts"].append(c2)
                    if c3:
                        current["actions"].append(c3)
                continue
            if current:
                sections.append(current)
            current = {"title": p, "facts": [], "actions": []}
            if c2:
                current["facts"].append(c2)
            if c3:
                current["actions"].append(c3)
        elif current:
            if c2:
                current["facts"].append(c2)
            if c3:
                current["actions"].append(c3)
        elif p:
            current = {"title": p, "facts": [c2] if c2 else [], "actions": [c3] if c3 else []}

    if current:
        sections.append(current)

    # Merge sample SKU into FBM if first section
    return sections


def li_items(items, cls=""):
    if not items:
        return ""
    tag = f' class="{cls}"' if cls else ""
    return "<ul" + tag + ">" + "".join(f"<li>{escape(x)}</li>" for x in items if x) + "</ul>"


def slide(title, body_html, slide_num=None, subtitle=""):
    num = f'<span class="slide-num">{slide_num}</span>' if slide_num else ""
    sub = f'<p class="slide-sub">{escape(subtitle)}</p>' if subtitle else ""
    return f"""
    <section class="slide">
      <header class="slide-head">
        <div>
          <h2>{escape(title)}</h2>
          {sub}
        </div>
        {num}
      </header>
      <div class="slide-body">{body_html}</div>
      <footer class="slide-foot">MetroShoe Warehouse · Master Sheet Data Analyst's Playbook</footer>
    </section>
    """


def main():
    sections = parse_playbook()
    all_tasks = []
    for s in sections:
        for a in s["actions"]:
            if a.lower().startswith("task:") or a.lower().startswith("question:"):
                all_tasks.append(f"<strong>{escape(s['title'])}</strong> — {escape(a)}")

    overview_items = "".join(f"<li><strong>{escape(s['title'])}</strong></li>" for s in sections)

    slides = [
        f"""
    <section class="slide slide-title">
      <div class="title-inner">
        <p class="eyebrow">MetroShoe Warehouse</p>
        <h1>Master Sheet Data Analyst's Playbook</h1>
        <p class="lead">Business logic and processes for Amazon marketplace operations — pricing, MAP, BuyBox, listings, and inventory decisions.</p>
        <p class="meta-line">Training reference · June 2026</p>
      </div>
      <footer class="slide-foot">Print: Ctrl+P → Save as PDF (one slide per page)</footer>
    </section>
        """,
        slide(
            "What this playbook covers",
            f"""
            <p>This guide explains <strong>why</strong> we manage the Master Sheet the way we do — not just the steps, but the business rules behind FBM SKUs, MAP pricing, BuyBox strategy, and listing health.</p>
            <div class="two-col">
              <div>
                <h3>Topics ({len(sections)})</h3>
                <ul class="compact">{overview_items}</ul>
              </div>
              <div>
                <h3>How to use it</h3>
                <ul>
                  <li>Read each topic's <strong>process notes</strong> (left column in the source sheet)</li>
                  <li>Follow <strong>action items</strong> and training tasks in orange callouts</li>
                  <li>Escalate grey areas to leadership (Emily, Zac, Rhett) when noted</li>
                </ul>
              </div>
            </div>
            """,
            slide_num="2",
        ),
    ]

    for i, s in enumerate(sections, start=3):
        facts_html = li_items(s["facts"])
        actions = [a for a in s["actions"] if a]
        actions_html = ""
        if actions:
            actions_html = '<div class="callout"><h3>Actions &amp; training notes</h3>' + li_items(actions) + "</div>"
        slides.append(slide(s["title"], facts_html + actions_html, slide_num=str(i)))

    slides.append(
        slide(
            "Training checklist",
            """
            <p>Complete these familiarization tasks as you work through the Master Sheet:</p>
            <ul class="checklist">
            """
            + "".join(
                f"<li>{t}</li>"
                for s in sections
                for a in s["actions"]
                if a.lower().startswith("task:") or a.lower().startswith("question:")
                for t in [f"<strong>{escape(s['title'])}</strong> — {escape(a)}"]
            )
            + """
            </ul>
            <div class="callout muted">
              <p><strong>Key contacts:</strong> Raise missing FBM SKUs to Emily. Clarify REMOVE status and MAP drops with Zac and Rhett. Listing issues → amazon@metroshoewarehouse.com</p>
            </div>
            """,
            slide_num=str(len(sections) + 3),
            subtitle="Onboarding & ongoing reference",
        )
    )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Master Sheet Data Analyst's Playbook — Presentation</title>
  <style>
    :root {{
      --ink: #1f2937;
      --muted: #4b5563;
      --accent: #ea580c;
      --accent-dark: #9a3412;
      --border: #e5e7eb;
      --bg: #f9fafb;
    }}
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{
      font-family: "Segoe UI", system-ui, -apple-system, sans-serif;
      color: var(--ink);
      background: #e5e7eb;
      line-height: 1.5;
    }}
    .slide {{
      width: 10in;
      min-height: 7.5in;
      margin: 24px auto;
      padding: 0.55in 0.65in 0.5in;
      background: #fff;
      box-shadow: 0 4px 24px rgba(0,0,0,.12);
      display: flex;
      flex-direction: column;
      page-break-after: always;
      break-after: page;
    }}
    .slide-title {{
      justify-content: center;
      background: linear-gradient(160deg, #1f2937 0%, #374151 45%, #ea580c 100%);
      color: #fff;
    }}
    .title-inner {{ max-width: 7.5in; }}
    .eyebrow {{
      text-transform: uppercase;
      letter-spacing: 0.12em;
      font-size: 10pt;
      opacity: 0.9;
      margin-bottom: 12px;
    }}
    .slide-title h1 {{
      font-size: 32pt;
      font-weight: 700;
      line-height: 1.15;
      margin-bottom: 16px;
    }}
    .lead {{
      font-size: 13pt;
      opacity: 0.95;
      max-width: 6.5in;
      margin-bottom: 24px;
    }}
    .meta-line {{ font-size: 10pt; opacity: 0.75; }}
    .slide-title .slide-foot {{ color: rgba(255,255,255,.65); border-color: rgba(255,255,255,.2); }}
    .slide-head {{
      display: flex;
      justify-content: space-between;
      align-items: flex-start;
      border-bottom: 3px solid var(--accent);
      padding-bottom: 10px;
      margin-bottom: 18px;
    }}
    .slide-head h2 {{
      font-size: 20pt;
      color: #111827;
      font-weight: 700;
    }}
    .slide-sub {{ font-size: 10pt; color: var(--muted); margin-top: 4px; }}
    .slide-num {{
      font-size: 28pt;
      font-weight: 700;
      color: #f3f4f6;
      line-height: 1;
    }}
    .slide-body {{ flex: 1; font-size: 11pt; }}
    .slide-body h3 {{
      font-size: 10pt;
      text-transform: uppercase;
      letter-spacing: 0.06em;
      color: var(--accent);
      margin: 14px 0 8px;
    }}
    .slide-body ul {{ padding-left: 22px; margin-bottom: 10px; }}
    .slide-body li {{ margin-bottom: 7px; }}
    ul.compact li {{ margin-bottom: 4px; }}
    .two-col {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 24px;
      margin-top: 12px;
    }}
    .callout {{
      background: #fff7ed;
      border: 1px solid #fdba74;
      border-radius: 8px;
      padding: 12px 14px;
      margin-top: 14px;
    }}
    .callout h3 {{ margin-top: 0; color: var(--accent-dark); }}
    .callout.muted {{ background: var(--bg); border-color: var(--border); }}
    .callout.muted p {{ color: var(--muted); font-size: 10pt; }}
    ul.checklist li {{ list-style: none; position: relative; padding-left: 22px; }}
    ul.checklist li::before {{
      content: "☐";
      position: absolute;
      left: 0;
      color: var(--accent);
    }}
    .slide-foot {{
      margin-top: auto;
      padding-top: 10px;
      border-top: 1px solid var(--border);
      font-size: 8.5pt;
      color: var(--muted);
    }}
    @media print {{
      body {{ background: #fff; }}
      .slide {{
        width: auto;
        min-height: auto;
        height: 100vh;
        margin: 0;
        box-shadow: none;
        page-break-after: always;
      }}
      @page {{ size: landscape; margin: 0.4in; }}
    }}
  </style>
</head>
<body>
{"".join(slides)}
</body>
</html>
"""

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(html, encoding="utf-8")
    print(f"Wrote {OUT} ({len(sections)} topic slides + title + overview + checklist)")


if __name__ == "__main__":
    main()
